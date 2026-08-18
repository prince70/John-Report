from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from typing import Optional
import pyodbc
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io

router = APIRouter()

DB_SERVER = "192.168.41.57"
DB_DATABASE = "department2020"
DB_USERNAME = "sa"
DB_PASSWORD = "3518i"

def get_db_connection():
    conn_str = (
        f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={DB_SERVER};DATABASE={DB_DATABASE};"
        f"UID={DB_USERNAME};PWD={DB_PASSWORD};"
    )
    return pyodbc.connect(conn_str)

def fetch_rows(sql, params):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute(sql, params)
        return cursor.fetchall()
    finally:
        try: conn.close()
        except: pass

def build_b(rows):
    result = []
    for row in rows:
        理论产能 = float(row[8]) if row[8] else 0
        result.append({
            "是否核对": row[0] or "", "序列号": row[2] or "", "生产车间": row[3] or "",
            "工单编号": row[4] or "", "工单状态": row[5] or "", "订单批号": row[6] or "",
            "订单数量": float(row[7]) if row[7] else 0, "料品编码": row[10] or "",
            "生产线编号": row[11] or "", "规格型号": row[12] or "", "报工人": row[13] or "",
            "工单数量": float(row[14]) if row[14] else 0, "报工数量": float(row[15]) if row[15] else 0,
            "报废数量": float(row[16]) if row[16] else 0, "返修数量": float(row[17]) if row[17] else 0,
            "车间提供": 理论产能,
            "时间": float(row[9]) if row[9] else (round((float(row[15] or 0) + float(row[17] or 0) + float(row[16] or 0)) / 理论产能, 2) if 理论产能 != 0 else None),
            "开工时间": str(row[18]) if row[18] else "", "完工时间": str(row[19]) if row[19] else "",
            "确定交期": row[20] or "", "工序": row[22] or "", "客户": row[21] or "",
            "急货": row[23] or "",
        })
    return result

build_a = build_b

def build_material(rows):
    result = []
    for row in rows:
        理论产能 = float(row[8]) if row[8] else 0
        result.append({
            "是否核对": row[0] or "", "序列号": row[2] or "", "生产车间": row[3] or "",
            "工单编号": row[4] or "", "工单状态": row[5] or "", "订单批号": row[6] or "",
            "订单数量": float(row[7]) if row[7] else 0, "料品编码": row[10] or "",
            "生产线编号": row[11] or "", "生产线描述": row[12] or "",
            "规格型号": row[13] or "", "成品料品名称": row[14] or "",
            "成品料品规格": row[15] or "",
            "报工人": row[16] or "", "工单数量": float(row[17]) if row[17] else 0,
            "报工数量": float(row[18]) if row[18] else 0,
            "报废数量": float(row[19]) if row[19] else 0,
            "返修数量": float(row[20]) if row[20] else 0,
            "车间提供": 理论产能,
            "时间": round((float(row[18] or 0) + float(row[20] or 0) + float(row[19] or 0)) / 理论产能, 2) if 理论产能 != 0 else None,
            "开工时间": str(row[21]) if row[21] else "", "完工时间": str(row[22]) if row[22] else "",
            "用料规格": row[23] or "",
        })
    return result

def build_zhuangqian(rows):
    result = []
    for row in rows:
        理论产能 = float(row[8]) if row[8] else 0
        result.append({
            "是否核对": row[0] or "", "序列号": row[2] or "", "生产车间": row[3] or "",
            "工单编号": row[4] or "", "工单状态": row[5] or "", "订单批号": row[6] or "",
            "订单数量": float(row[7]) if row[7] else 0, "料品编码": row[10] or "",
            "生产线编号": row[11] or "", "料品名称": row[12] or "", "规格型号": row[13] or "",
            "报工人": row[14] or "", "工单数量": float(row[15]) if row[15] else 0,
            "报工数量": float(row[16]) if row[16] else 0, "报废数量": float(row[17]) if row[17] else 0,
            "返修数量": float(row[18]) if row[18] else 0,
            "车间提供": 理论产能,
            "时间": float(row[9]) if row[9] else None,
            "开工时间": str(row[19]) if row[19] else "", "完工时间": str(row[20]) if row[20] else "",
            "确定交期": row[21] or "", "客户": row[22] or "", "工序": row[23] or "",
        })
    return result

def build_suoliang(rows):
    result = []
    for row in rows:
        理论产能 = float(row[8]) if row[8] else 0
        result.append({
            "是否核对": row[0] or "", "序列号": row[2] or "", "生产车间": row[3] or "",
            "工单编号": row[4] or "", "工单状态": row[5] or "", "订单批号": row[6] or "",
            "订单数量": float(row[7]) if row[7] else 0, "料品编码": row[10] or "",
            "生产线编号": row[11] or "", "规格型号": row[12] or "", "报工人": row[13] or "",
            "工单数量": float(row[14]) if row[14] else 0, "报工数量": float(row[15]) if row[15] else 0,
            "报工重量": float(row[16]) if row[16] else 0,
            "报废数量": float(row[17]) if row[17] else 0, "返修数量": float(row[18]) if row[18] else 0,
            "车间提供": 理论产能,
            "时间": float(row[9]) if row[9] else None,
            "开工时间": str(row[19]) if row[19] else "", "完工时间": str(row[20]) if row[20] else "",
            "确定交期": row[21] or "", "客户": row[22] or "", "工序": row[23] or "",
            "急货": row[24] or "",
        })
    return result

def build_c(rows):
    result = []
    for row in rows:
        理论产能 = float(row[8]) if row[8] else 0
        result.append({
            "是否核对": row[0] or "", "序列号": row[2] or "", "生产车间": row[3] or "",
            "工单编号": row[4] or "", "工单状态": row[5] or "", "订单批号": row[6] or "",
            "订单数量": float(row[7]) if row[7] else 0, "料品编码": row[10] or "",
            "生产线编号": row[11] or "", "规格型号": row[12] or "", "报工人": row[13] or "",
            "工单数量": float(row[14]) if row[14] else 0, "报工数量": float(row[15]) if row[15] else 0,
            "报废数量": float(row[16]) if row[16] else 0, "返修数量": float(row[17]) if row[17] else 0,
            "车间提供": 理论产能,
            "时间": float(row[9]) if row[9] else None,
            "开工时间": str(row[18]) if row[18] else "", "完工时间": str(row[19]) if row[19] else "",
            "确定交期": row[20] or "", "工序": row[21] or "", "客户": row[22] or "",
            "急货": row[23] or "",
        })
    return result

def build_d(rows):
    result = []
    for row in rows:
        理论产能 = float(row[8]) if row[8] else 0
        result.append({
            "是否核对": row[0] or "", "序列号": row[2] or "", "生产车间": row[3] or "",
            "工单编号": row[4] or "", "工单状态": row[5] or "", "订单批号": row[6] or "",
            "订单数量": float(row[7]) if row[7] else 0, "料品编码": row[10] or "",
            "生产线编号": row[11] or "", "规格型号": row[12] or "", "报工人": row[13] or "",
            "工单数量": float(row[14]) if row[14] else 0, "报工数量": float(row[15]) if row[15] else 0,
            "报废数量": float(row[17]) if row[17] else 0, "返修数量": float(row[18]) if row[18] else 0,
            "车间提供": 理论产能,
            "时间": round((float(row[15] or 0) + float(row[18] or 0) + float(row[17] or 0)) / 理论产能, 2) if 理论产能 != 0 else None,
            "开工时间": str(row[19]) if row[19] else "", "完工时间": str(row[20]) if row[20] else "",
            "确定交期": row[21] or "", "工序": row[23] or "", "客户": row[22] or "",
            "急货": row[24] or "",
        })
    return result

SQL_A = """
SELECT a.isCheck, a.iNo, b._Identify, b.生产车间,
a.JobExternalId AS 工单编号, b.工单状态, a.OrderNumber AS 订单批号, b.订单数量,
cap.理论产能,
CAST((ISNULL(a.EachFinishedQty, 0) + ISNULL(a.repairQty, 0) + ISNULL(a.scrapQty, 0))
    / NULLIF(cap.理论产能, 0) AS DECIMAL(18, 1)) AS 理论工时,
a.ItemExternalId AS 料品编码, a.ResName AS 生产线编号, a.ProductDescription AS 规格型号,
a.emp_name AS 报工人, b.计划产量 AS 工单数量, a.EachFinishedQty AS 报工数量,
a.scrapQty AS 报废数量, a.repairQty AS 返修数量,
a.StartDate AS 开工时间, a.FinishedDate AS 完工时间, b.确定交期,
b.OpExternalId AS 工序, b.客户, a.urgent
FROM APS_FinishedQty_ST a
JOIN 派工单 b ON a.JobExternalId = b.工单编号
LEFT JOIN APS.APS_SUO.dbo.ProductRules pr
    ON a.ResName = pr.ResourceExternalId AND a.ItemExternalId = pr.ProductItemExternalId
LEFT JOIN APS.APS_SUO.dbo.offline_process d
    ON b.OpExternalId = d.proccess AND a.ItemExternalId = d.item_no
CROSS APPLY (
    SELECT CASE WHEN pr.QtyPerCycle IS NOT NULL THEN pr.QtyPerCycle * 60.0
                WHEN d.capacity IS NOT NULL THEN CAST(d.capacity AS FLOAT) ELSE NULL END AS 理论产能
) AS cap
WHERE (
    (a.EachFinishedQty >= 0 AND b.生产车间 LIKE '锁体A%')
    OR (a.EachFinishedQty >= 0 AND b.生产车间 LIKE '锁体B%' AND b.OpExternalId LIKE '锁体A%')
    OR (a.EachFinishedQty >= 0 AND b.生产车间 LIKE '锁体C%' AND b.OpExternalId LIKE '锁体A%')
)
AND NOT (a.EachFinishedQty >= 0 AND b.生产车间 LIKE '锁体A%' AND b.OpExternalId LIKE '%锁体B%')
AND a.FinishedDate between ? and ?
"""

COUNT_A = """
SELECT COUNT(*) FROM APS_FinishedQty_ST a
JOIN 派工单 b ON a.JobExternalId = b.工单编号
WHERE (
    (a.EachFinishedQty >= 0 AND b.生产车间 LIKE '锁体A%')
    OR (a.EachFinishedQty >= 0 AND b.生产车间 LIKE '锁体B%' AND b.OpExternalId LIKE '锁体A%')
    OR (a.EachFinishedQty >= 0 AND b.生产车间 LIKE '锁体C%' AND b.OpExternalId LIKE '锁体A%')
)
AND NOT (a.EachFinishedQty >= 0 AND b.生产车间 LIKE '锁体A%' AND b.OpExternalId LIKE '%锁体B%')
AND a.FinishedDate between ? and ?
"""

SQL_SUOLIANG = """
select a.isCheck, a.iNo, b._Identify, b.生产车间, a.JobExternalId as 工单编号, b.工单状态,
a.OrderNumber as 订单批号, b.订单数量,
(60*pr.QtyPerCycle) as 理论产能,
CAST((isnull(a.EachFinishedQty,0)+isnull(a.repairQty,0)+isnull(a.scrapQty,0))/(60*pr.QtyPerCycle) AS DECIMAL(18,1)) as 理论工时,
a.ItemExternalId as 料品编码, a.ResName as 生产线编号, a.ProductDescription as 规格型号,
a.emp_name as 报工人, b.计划产量 as 工单数量, a.EachFinishedQty as 报工数量,
a.EachFinishedQtykg as 报工重量, a.scrapQty as 报废数量, a.repairQty as 返修数量,
a.StartDate as 开工时间, a.FinishedDate as 完工时间, b.确定交期, b.客户, b.OpExternalId, a.urgent
from APS_FinishedQty_SL a
join 派工单 b on a.JobExternalId = b.工单编号
left join APS.APS_SUO.dbo.ProductRules pr
    on a.ResName = pr.ResourceExternalId and a.ItemExternalId = pr.ProductItemExternalId
where a.EachFinishedQty > 0 and b.生产车间 = '锁梁车间'
and a.FinishedDate between ? and ?
"""

COUNT_SUOLIANG = """
SELECT COUNT(*) FROM APS_FinishedQty_SL a
JOIN 派工单 b ON a.JobExternalId = b.工单编号
WHERE a.EachFinishedQty > 0 AND b.生产车间 = '锁梁车间'
AND a.FinishedDate between ? and ?
"""

SQL_B = """
select a.isCheck, a.iNo, b._Identify, b.生产车间, a.JobExternalId as 工单编号, b.工单状态,
a.OrderNumber as 订单批号, b.订单数量,
(60*pr.QtyPerCycle) as 理论产能,
CAST((isnull(a.EachFinishedQty,0)+isnull(a.repairQty,0)+isnull(a.scrapQty,0))/(60*pr.QtyPerCycle) AS DECIMAL(18,1)) as 理论工时,
a.ItemExternalId as 料品编码, a.ResName as 生产线编号, a.ProductDescription as 规格型号,
a.emp_name as 报工人, b.计划产量 as 工单数量, a.EachFinishedQty as 报工数量,
a.scrapQty as 报废数量, a.repairQty as 返修数量, b.确定交期, a.StartDate as 开工时间,
a.FinishedDate as 完工时间, b.客户, b.OpExternalId, a.urgent
from APS_FinishedQty_ST a
join 派工单 b on a.JobExternalId = b.工单编号
left join APS.APS_SUO.dbo.ProductRules pr
    on a.ResName = pr.ResourceExternalId and a.ItemExternalId = pr.ProductItemExternalId
where ((a.EachFinishedQty >=0 and b.生产车间 like '锁体B%')
  and NOT (a.EachFinishedQty >=0 and b.生产车间 like '锁体B%' and b.OpExternalId like '锁体A%')
  or (a.EachFinishedQty >=0 and b.生产车间 like '锁体A%' and b.OpExternalId like '%锁体B%'))
and a.FinishedDate between ? and ?
"""

COUNT_B = """
SELECT COUNT(*) FROM APS_FinishedQty_ST a
JOIN 派工单 b ON a.JobExternalId = b.工单编号
WHERE ((a.EachFinishedQty >=0 AND b.生产车间 LIKE '锁体B%')
  AND NOT (a.EachFinishedQty >=0 AND b.生产车间 LIKE '锁体B%' AND b.OpExternalId LIKE '锁体A%')
  OR (a.EachFinishedQty >=0 AND b.生产车间 LIKE '锁体A%' AND b.OpExternalId LIKE '%锁体B%'))
AND a.FinishedDate between ? and ?
"""

SQL_C = """
SELECT a.isCheck, a.iNo, b._Identify, b.生产车间,
a.JobExternalId AS 工单编号, b.工单状态, a.OrderNumber AS 订单批号, b.订单数量,
CASE WHEN pr.QtyPerCycle IS NOT NULL THEN pr.QtyPerCycle * 60.0
     WHEN d.capacity IS NOT NULL THEN CAST(d.capacity AS FLOAT) ELSE NULL END AS 理论产能,
CAST((ISNULL(a.EachFinishedQty, 0) + ISNULL(a.repairQty, 0) + ISNULL(a.scrapQty, 0))
    / NULLIF(CASE WHEN pr.QtyPerCycle IS NOT NULL THEN pr.QtyPerCycle * 60.0
                  WHEN d.capacity IS NOT NULL THEN CAST(d.capacity AS FLOAT) ELSE NULL END, 0)
    AS DECIMAL(18, 1)) AS 理论工时,
a.ItemExternalId AS 料品编码, a.ResName AS 生产线编号, a.ProductDescription AS 规格型号,
a.emp_name AS 报工人, b.计划产量 AS 工单数量, a.EachFinishedQty AS 报工数量,
a.scrapQty AS 报废数量, a.repairQty AS 返修数量,
a.StartDate AS 开工时间, a.FinishedDate AS 完工时间, b.确定交期,
b.OpExternalId AS 工序, b.客户, a.urgent
FROM APS_FinishedQty_ST a
JOIN 派工单 b ON a.JobExternalId = b.工单编号
LEFT JOIN APS.APS_SUO.dbo.ProductRules pr
    ON a.ResName = pr.ResourceExternalId AND a.ItemExternalId = pr.ProductItemExternalId
LEFT JOIN APS.APS_SUO.dbo.offline_process d
    ON b.OpExternalId = d.proccess AND a.ItemExternalId = d.item_no
WHERE a.EachFinishedQty >= 0 AND b.生产车间 LIKE '锁体C%'
AND a.FinishedDate between ? and ?
"""

COUNT_C = """
SELECT COUNT(*) FROM APS_FinishedQty_ST a
JOIN 派工单 b ON a.JobExternalId = b.工单编号
WHERE a.EachFinishedQty >= 0 AND b.生产车间 LIKE '锁体C%'
AND a.FinishedDate between ? and ?
"""

SQL_D = """
select a.isCheck, a.iNo, b._Identify, b.生产车间, a.JobExternalId as 工单编号, b.工单状态,
a.OrderNumber as 订单批号, b.订单数量,
(60*pr.QtyPerCycle) as 理论产能,
CAST((isnull(a.EachFinishedQty,0)+isnull(a.repairQty,0)+isnull(a.scrapQty,0))/(60*pr.QtyPerCycle) AS DECIMAL(18,1)) as 理论工时,
a.ItemExternalId as 料品编码, a.ResName as 生产线编号, a.ProductDescription as 规格型号,
a.emp_name as 报工人, b.计划产量 as 工单数量, a.EachFinishedQty as 报工数量,
a.EachFinishedQtykg as 报工重量, a.scrapQty as 报废数量, a.repairQty as 返修数量,
a.StartDate as 开工时间, a.FinishedDate as 完工时间, b.确定交期, b.客户, b.OpExternalId, a.urgent
from APS_FinishedQty_ST a
join 派工单 b on a.JobExternalId = b.工单编号
left join APS.APS_SUO.dbo.ProductRules pr
    on a.ResName = pr.ResourceExternalId and a.ItemExternalId = pr.ProductItemExternalId
where a.EachFinishedQty >=0 and b.生产车间 like '锁体D%'
and a.FinishedDate between ? and ?
"""

COUNT_D = """
SELECT COUNT(*) FROM APS_FinishedQty_ST a
JOIN 派工单 b ON a.JobExternalId = b.工单编号
WHERE a.EachFinishedQty >=0 AND b.生产车间 LIKE '锁体D%'
AND a.FinishedDate between ? and ?
"""

SQL_MATERIAL = """
select a.isCheck, a.iNo, b._Identify, b.生产车间, a.JobExternalId as 工单编号, b.工单状态,
a.OrderNumber as 订单批号, b.订单数量,
(60*pr.QtyPerCycle) as 理论产能,
CAST((isnull(a.EachFinishedQty,0)+isnull(a.repairQty,0)+isnull(a.scrapQty,0))/(60*pr.QtyPerCycle) AS DECIMAL(18,1)) as 理论工时,
a.ItemExternalId as 料品编码, a.ResName as 生产线编号,
rc.Description as 生产线描述,
a.ProductDescription as 规格型号,
c.part_name as 成品料品名称,
c.part_spec as 成品料品规格,
a.emp_name as 报工人, b.计划产量 as 工单数量, a.EachFinishedQty as 报工数量,
a.scrapQty as 报废数量, a.repairQty as 返修数量,
a.StartDate as 开工时间, a.FinishedDate as 完工时间, b.general_name
from APS_FinishedQty_ST a
join 派工单 b on a.JobExternalId = b.工单编号
left join APS.APS_SUO.dbo.ProductRules pr
    on a.ResName = pr.ResourceExternalId and a.ItemExternalId = pr.ProductItemExternalId
left join [V_销售订单] c on b.订单批号 = c.sheet_lot
left join APS.APS_SUO.dbo.Resource rc on a.ResName = rc.ExternalId
where a.EachFinishedQty > 0 and b.生产车间 = '开料车间'
and a.FinishedDate between ? and ?
"""

COUNT_MATERIAL = """
SELECT COUNT(*) FROM APS_FinishedQty_ST a
JOIN 派工单 b ON a.JobExternalId = b.工单编号
LEFT JOIN APS.APS_SUO.dbo.Resource rc ON a.ResName = rc.ExternalId
WHERE a.EachFinishedQty > 0 AND b.生产车间 = '开料车间'
AND a.FinishedDate between ? and ?
"""

SQL_ZHUANGQIAN = """
select a.isCheck, a.iNo, b._Identify, b.生产车间, a.JobExternalId as 工单编号, b.工单状态,
a.OrderNumber as 订单批号, b.订单数量,
(60*pr.QtyPerCycle) as 理论产能,
CAST((isnull(a.EachFinishedQty,0)+isnull(a.repairQty,0)+isnull(a.scrapQty,0))/(60*pr.QtyPerCycle) AS DECIMAL(18,1)) as 理论工时,
a.ItemExternalId as 料品编码, a.ResName as 生产线编号, b.料品名称, a.ProductDescription as 规格型号,
a.emp_name as 报工人, b.计划产量 as 工单数量, a.EachFinishedQty as 报工数量,
a.scrapQty as 报废数量, a.repairQty as 返修数量,
a.StartDate as 开工时间, a.FinishedDate as 完工时间, b.确定交期, b.客户, b.OpExternalId
from APS_FinishedQty a
join 派工单 b on a.JobExternalId = b.工单编号
left join APS.APS_SUO.dbo.ProductRules pr
    on a.ResName = pr.ResourceExternalId and a.ItemExternalId = pr.ProductItemExternalId
where a.EachFinishedQty >= 0 and b.锁类分区 in ('铝门锁区','胆仔锁区','功能锁区')
and a.FinishedDate between ? and ?
"""

COUNT_ZHUANGQIAN = """
SELECT COUNT(*) FROM APS_FinishedQty a
JOIN 派工单 b ON a.JobExternalId = b.工单编号
WHERE a.EachFinishedQty >= 0 AND b.锁类分区 IN ('铝门锁区','胆仔锁区','功能锁区')
AND a.FinishedDate between ? and ?
"""

def build_damo(rows):
    result = []
    for row in rows:
        理论产能 = float(row[9]) if row[9] else 0
        result.append({
            "是否核对": row[0] or "", "序列号": row[2] or "", "生产车间": row[3] or "",
            "工单编号": row[4] or "", "工单状态": row[5] or "", "订单批号": row[6] or "",
            "确定交期": row[7] or "", "订单数量": float(row[8]) if row[8] else 0,
            "料品编码": row[11] or "", "生产线编号": row[12] or "",
            "规格型号": row[13] or "", "报工人": row[14] or "",
            "工单数量": float(row[15]) if row[15] else 0, "报工数量": float(row[16]) if row[16] else 0,
            "报废数量": float(row[17]) if row[17] else 0, "返修数量": float(row[18]) if row[18] else 0,
            "车间提供": 理论产能,
            "时间": round((float(row[16] or 0) + float(row[18] or 0) + float(row[17] or 0)) / 理论产能, 2) if 理论产能 != 0 else None,
            "开工时间": str(row[19]) if row[19] else "", "完工时间": str(row[20]) if row[20] else "",
            "客户": row[21] or "", "工序": row[22] or "",
        })
    return result

def build_damo_dm(rows):
    result = []
    for row in rows:
        理论产能 = float(row[9]) if row[9] else 0
        result.append({
            "是否核对": row[0] or "", "序列号": row[2] or "", "生产车间": row[3] or "",
            "工单编号": row[4] or "", "工单状态": row[5] or "", "订单批号": row[6] or "",
            "确定交期": row[7] or "", "订单数量": float(row[8]) if row[8] else 0,
            "料品编码": row[11] or "", "料品名称": row[12] or "",
            "生产线编号": row[13] or "", "规格型号": row[14] or "",
            "报工人": row[15] or "", "工单数量": float(row[16]) if row[16] else 0,
            "报工数量": float(row[17]) if row[17] else 0, "报废数量": float(row[18]) if row[18] else 0,
            "返修数量": float(row[19]) if row[19] else 0,
            "车间提供": 理论产能,
            "时间": round((float(row[17] or 0) + float(row[19] or 0) + float(row[18] or 0)) / 理论产能, 2) if 理论产能 != 0 else None,
            "开工时间": str(row[20]) if row[20] else "", "完工时间": str(row[21]) if row[21] else "",
            "客户": row[22] or "", "工序": row[23] or "",
            "急货": row[24] or "",
        })
    return result

SQL_DAMO = """
select a.isCheck, a.iNo, b._Identify, b.生产车间, a.JobExternalId as 工单编号, b.工单状态,
a.OrderNumber as 订单批号, b.确定交期, b.订单数量,
(60*pr.QtyPerCycle) as 理论产能,
CAST((isnull(a.EachFinishedQty,0)+isnull(a.repairQty,0)+isnull(a.scrapQty,0))/(60*pr.QtyPerCycle) AS DECIMAL(18,1)) as 理论工时,
a.ItemExternalId as 料品编码, a.ResName as 生产线编号, a.ProductDescription as 规格型号,
a.emp_name as 报工人, b.计划产量 as 工单数量, a.EachFinishedQty as 报工数量,
a.scrapQty as 报废数量, a.repairQty as 返修数量,
a.StartDate as 开工时间, a.FinishedDate as 完工时间, b.客户, b.OpExternalId
from APS_FinishedQty a
join 派工单 b on a.JobExternalId = b.工单编号
left join APS.APS_SUO.dbo.ProductRules pr
    on a.ResName = pr.ResourceExternalId and a.ItemExternalId = pr.ProductItemExternalId
where a.EachFinishedQty > 0 and b.锁类分区 in ('普通挂锁区')
and a.FinishedDate between ? and ?
"""

COUNT_DAMO = """
SELECT COUNT(*) FROM APS_FinishedQty a
JOIN 派工单 b ON a.JobExternalId = b.工单编号
WHERE a.EachFinishedQty > 0 AND b.锁类分区 IN ('普通挂锁区')
AND a.FinishedDate between ? and ?
"""

SQL_DAMO_DM = """
select a.isCheck, a.iNo, b._Identify, b.生产车间, a.JobExternalId as 工单编号, b.工单状态,
a.OrderNumber as 订单批号, b.确定交期, b.订单数量,
(60*pr.QtyPerCycle) as 理论产能,
CAST((isnull(a.EachFinishedQty,0)+isnull(a.repairQty,0)+isnull(a.scrapQty,0))/(60*pr.QtyPerCycle) AS DECIMAL(18,1)) as 理论工时,
a.ItemExternalId as 料品编码, b.料品名称, a.ResName as 生产线编号, a.ProductDescription as 规格型号,
a.emp_name as 报工人, b.计划产量 as 工单数量, a.EachFinishedQty as 报工数量,
a.scrapQty as 报废数量, a.repairQty as 返修数量,
a.StartDate as 开工时间, a.FinishedDate as 完工时间, b.客户, b.OpExternalId, a.urgent
from APS_FinishedQty_DM a
join 派工单 b on a.JobExternalId = b.工单编号
left join APS.APS_SUO.dbo.ProductRules pr
    on a.ResName = pr.ResourceExternalId and a.ItemExternalId = pr.ProductItemExternalId
where a.EachFinishedQty > 0 and b.生产车间 = '打磨车间-打磨区'
and a.FinishedDate between ? and ?
"""

COUNT_DAMO_DM = """
SELECT COUNT(*) FROM APS_FinishedQty_DM a
JOIN 派工单 b ON a.JobExternalId = b.工单编号
WHERE a.EachFinishedQty > 0 AND b.生产车间 = '打磨车间-打磨区'
AND a.FinishedDate between ? and ?
"""

def _paginated_query(base_sql, params, build_fn, page=1, page_size=100, order_by="b.订单批号", count_sql=None):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        if count_sql is None:
            count_sql = f"SELECT COUNT(*) FROM ({base_sql}) AS _t"
        cursor.execute(count_sql, list(params))
        total_count = cursor.fetchone()[0]
        offset = (page - 1) * page_size
        data_sql = f"{base_sql} ORDER BY {order_by} OFFSET ? ROWS FETCH NEXT ? ROWS ONLY"
        data_params = list(params) + [offset, page_size]
        cursor.execute(data_sql, data_params)
        rows = cursor.fetchall()
        data = build_fn(rows)
        return {"status": "success", "data": data, "total_count": total_count}
    finally:
        try: conn.close()
        except: pass

def _query_damo_dm(start, end, 序列号, 姓名, 规格型号, 工序, 生产线编号, 工单批号, page=1, page_size=100):
    try:
        extra = ""
        params = [start, end]
        if 序列号:
            extra += " AND b._Identify LIKE ?"
            params.append(f"%{序列号}%")
        if 姓名:
            extra += " AND a.emp_name LIKE ?"
            params.append(f"%{姓名}%")
        if 规格型号:
            extra += " AND a.ProductDescription LIKE ?"
            params.append(f"%{规格型号}%")
        if 工序:
            extra += " AND b.OpExternalId LIKE ?"
            params.append(f"%{工序}%")
        if 生产线编号:
            extra += " AND a.ResName LIKE ?"
            params.append(f"%{生产线编号}%")
        if 工单批号:
            extra += " AND a.OrderNumber LIKE ?"
            params.append(f"%{工单批号}%")
        return _paginated_query(SQL_DAMO_DM + extra, params, build_damo_dm, page, page_size, count_sql=COUNT_DAMO_DM + extra)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"服务器错误: {exc}")

def _query_damo(start, end, 序列号, 姓名, 生产线编号, 工单批号, page=1, page_size=100):
    try:
        extra = ""
        params = [start, end]
        if 序列号:
            extra += " AND b._Identify LIKE ?"
            params.append(f"%{序列号}%")
        if 姓名:
            extra += " AND a.emp_name LIKE ?"
            params.append(f"%{姓名}%")
        if 生产线编号:
            extra += " AND a.ResName LIKE ?"
            params.append(f"%{生产线编号}%")
        if 工单批号:
            extra += " AND a.OrderNumber LIKE ?"
            params.append(f"%{工单批号}%")
        return _paginated_query(SQL_DAMO + extra, params, build_damo, page, page_size, count_sql=COUNT_DAMO + extra)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"服务器错误: {exc}")

def _query_zhuangqian(start, end, 序列号, 姓名, 工单批号, 锁类分区, 含TSA, page=1, page_size=100):
    try:
        extra = ""
        params = [start, end]
        if 序列号:
            extra += " AND b._Identify LIKE ?"
            params.append(f"%{序列号}%")
        if 姓名:
            extra += " AND a.emp_name LIKE ?"
            params.append(f"%{姓名}%")
        if 工单批号:
            extra += " AND a.OrderNumber LIKE ?"
            params.append(f"%{工单批号}%")
        if 锁类分区:
            extra += " AND b.锁类分区 LIKE ?"
            params.append(f"%{锁类分区}%")
        if not 含TSA:
            extra += " AND b.料品名称 <> 'TSA006'"
        return _paginated_query(SQL_ZHUANGQIAN + extra, params, build_zhuangqian, page, page_size, count_sql=COUNT_ZHUANGQIAN + extra)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"服务器错误: {exc}")

def _query_workshop(sql, build_fn, start, end, 序列号, 姓名, 订单批号, page=1, page_size=100, count_sql=None):
    try:
        extra = ""
        params = [start, end]
        if 序列号:
            extra += " AND b._Identify LIKE ?"
            params.append(f"%{序列号}%")
        if 姓名:
            extra += " AND a.emp_name LIKE ?"
            params.append(f"%{姓名}%")
        if 订单批号:
            extra += " AND a.OrderNumber LIKE ?"
            params.append(f"%{订单批号}%")
        final_count_sql = (count_sql + extra) if count_sql else None
        return _paginated_query(sql + extra, params, build_fn, page, page_size, count_sql=final_count_sql)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"服务器错误: {exc}")

def _query_material(sql, build_fn, start, end, 序列号, 姓名, 工单编号, 生产线编号=None, 生产线描述=None, page=1, page_size=100, count_sql=None):
    try:
        extra = ""
        count_extra = ""
        params = [start, end]
        if 序列号:
            extra += " AND b._Identify LIKE ?"
            count_extra += " AND b._Identify LIKE ?"
            params.append(f"%{序列号}%")
        if 姓名:
            extra += " AND a.emp_name LIKE ?"
            count_extra += " AND a.emp_name LIKE ?"
            params.append(f"%{姓名}%")
        if 工单编号:
            extra += " AND a.JobExternalId LIKE ?"
            count_extra += " AND a.JobExternalId LIKE ?"
            params.append(f"%{工单编号}%")
        if 生产线编号:
            extra += " AND a.ResName LIKE ?"
            count_extra += " AND a.ResName LIKE ?"
            params.append(f"%{生产线编号}%")
        if 生产线描述:
            extra += " AND rc.Description LIKE ?"
            count_extra += " AND rc.Description LIKE ?"
            params.append(f"%{生产线描述}%")
        final_count_sql = (count_sql + count_extra) if count_sql else None
        return _paginated_query(sql + extra, params, build_fn, page, page_size, count_sql=final_count_sql)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"服务器错误: {exc}")

def _get_material_options(field):
    try:
        if field == "生产线编号":
            sql = "SELECT DISTINCT a.ResName FROM APS_FinishedQty_ST a JOIN 派工单 b ON a.JobExternalId = b.工单编号 WHERE a.EachFinishedQty > 0 AND b.生产车间 = '开料车间' AND a.ResName IS NOT NULL AND a.ResName != '' ORDER BY a.ResName"
        elif field == "生产线描述":
            sql = "SELECT DISTINCT rc.Description FROM APS_FinishedQty_ST a JOIN 派工单 b ON a.JobExternalId = b.工单编号 LEFT JOIN APS.APS_SUO.dbo.Resource rc ON a.ResName = rc.ExternalId WHERE a.EachFinishedQty > 0 AND b.生产车间 = '开料车间' AND rc.Description IS NOT NULL AND rc.Description != '' ORDER BY rc.Description"
        else:
            return []
        rows = fetch_rows(sql, [])
        return [str(row[0]) for row in rows]
    except:
        return []

@router.get("/workshopReportDetail/LockB", summary="锁体B报工详情")
async def lock_b(start: str = Query(...), end: str = Query(...), 序列号: Optional[str] = None, 姓名: Optional[str] = None, 订单批号: Optional[str] = None, page: int = 1, page_size: int = 100):
    return _query_workshop(SQL_B, build_b, start, end, 序列号, 姓名, 订单批号, page, page_size, count_sql=COUNT_B)

@router.get("/workshopReportDetail/LockC", summary="锁体C报工详情")
async def lock_c(start: str = Query(...), end: str = Query(...), 序列号: Optional[str] = None, 姓名: Optional[str] = None, 订单批号: Optional[str] = None, page: int = 1, page_size: int = 100):
    return _query_workshop(SQL_C, build_c, start, end, 序列号, 姓名, 订单批号, page, page_size, count_sql=COUNT_C)

@router.get("/workshopReportDetail/LockD", summary="锁体D报工详情")
async def lock_d(start: str = Query(...), end: str = Query(...), 序列号: Optional[str] = None, 姓名: Optional[str] = None, 订单批号: Optional[str] = None, page: int = 1, page_size: int = 100):
    return _query_workshop(SQL_D, build_d, start, end, 序列号, 姓名, 订单批号, page, page_size, count_sql=COUNT_D)

@router.get("/workshopReportDetail/LockA", summary="锁体A报工详情")
async def lock_a(start: str = Query(...), end: str = Query(...), 序列号: Optional[str] = None, 姓名: Optional[str] = None, 订单批号: Optional[str] = None, page: int = 1, page_size: int = 100):
    return _query_workshop(SQL_A, build_a, start, end, 序列号, 姓名, 订单批号, page, page_size, count_sql=COUNT_A)

@router.get("/workshopReportDetail/Suoliang", summary="锁梁报工详情")
async def suoliang(start: str = Query(...), end: str = Query(...), 序列号: Optional[str] = None, 姓名: Optional[str] = None, 订单批号: Optional[str] = None, page: int = 1, page_size: int = 100):
    return _query_workshop(SQL_SUOLIANG, build_suoliang, start, end, 序列号, 姓名, 订单批号, page, page_size, count_sql=COUNT_SUOLIANG)

@router.get("/workshopReportDetail/Material", summary="开料车间报工详情")
async def material(start: str = Query(...), end: str = Query(...), 工单编号: Optional[str] = None, 序列号: Optional[str] = None, 姓名: Optional[str] = None, 生产线编号: Optional[str] = None, 生产线描述: Optional[str] = None, page: int = 1, page_size: int = 100):
    return _query_material(SQL_MATERIAL, build_material, start, end, 序列号, 姓名, 工单编号, 生产线编号, 生产线描述, page, page_size, count_sql=COUNT_MATERIAL)

@router.get("/workshopReportDetail/Material/options", summary="开料车间下拉选项")
async def material_options(field: str = Query(...)):
    return {"status": "success", "data": _get_material_options(field)}

@router.get("/workshopReportDetail/Zhuangqian", summary="装嵌报工详情")
async def zhuangqian(start: str = Query(...), end: str = Query(...), 序列号: Optional[str] = None, 姓名: Optional[str] = None, 工单批号: Optional[str] = None, 锁类分区: Optional[str] = None, 含TSA: bool = False, page: int = 1, page_size: int = 100):
    return _query_zhuangqian(start, end, 序列号, 姓名, 工单批号, 锁类分区, 含TSA, page, page_size)

@router.get("/workshopReportDetail/Damo", summary="打磨-装配区报工详情")
async def damo(start: str = Query(...), end: str = Query(...), 序列号: Optional[str] = None, 姓名: Optional[str] = None, 生产线编号: Optional[str] = None, 工单批号: Optional[str] = None, page: int = 1, page_size: int = 100):
    return _query_damo(start, end, 序列号, 姓名, 生产线编号, 工单批号, page, page_size)

@router.get("/workshopReportDetail/DamoDM", summary="打磨-打磨区报工详情")
async def damo_dm(start: str = Query(...), end: str = Query(...), 序列号: Optional[str] = None, 姓名: Optional[str] = None, 规格型号: Optional[str] = None, 工序: Optional[str] = None, 生产线编号: Optional[str] = None, 工单批号: Optional[str] = None, page: int = 1, page_size: int = 100):
    return _query_damo_dm(start, end, 序列号, 姓名, 规格型号, 工序, 生产线编号, 工单批号, page, page_size)

# --- 导出功能 ---

EXPORT_COLUMNS = {
    "lock_a": ["是否核对","序列号","生产车间","工单编号","工单状态","订单批号","订单数量","料品编码","生产线编号","规格型号","报工人","工单数量","报工数量","报废数量","返修数量","车间提供","时间","开工时间","完工时间","确定交期","工序","客户","急货"],
    "lock_b": ["是否核对","序列号","生产车间","工单编号","工单状态","订单批号","订单数量","料品编码","生产线编号","规格型号","报工人","工单数量","报工数量","报废数量","返修数量","车间提供","时间","开工时间","完工时间","确定交期","工序","客户","急货"],
    "lock_c": ["是否核对","序列号","生产车间","工单编号","工单状态","订单批号","订单数量","料品编码","生产线编号","规格型号","报工人","工单数量","报工数量","报废数量","返修数量","车间提供","时间","开工时间","完工时间","确定交期","工序","客户","急货"],
    "lock_d": ["是否核对","序列号","生产车间","工单编号","工单状态","订单批号","订单数量","料品编码","生产线编号","规格型号","报工人","工单数量","报工数量","报废数量","返修数量","车间提供","时间","开工时间","完工时间","确定交期","工序","客户","急货"],
    "suoliang": ["是否核对","序列号","生产车间","工单编号","工单状态","订单批号","订单数量","料品编码","生产线编号","规格型号","报工人","工单数量","报工数量","报工重量","报废数量","返修数量","车间提供","时间","开工时间","完工时间","确定交期","客户","工序","急货"],
    "material": ["是否核对","序列号","生产车间","工单编号","工单状态","订单批号","订单数量","料品编码","生产线编号","生产线描述","规格型号","成品料品名称","成品料品规格","报工人","工单数量","报工数量","报废数量","返修数量","车间提供","时间","开工时间","完工时间","用料规格"],
    "zhuangqian": ["是否核对","序列号","生产车间","工单编号","工单状态","订单批号","订单数量","料品编码","生产线编号","料品名称","规格型号","报工人","工单数量","报工数量","报废数量","返修数量","车间提供","时间","开工时间","完工时间","确定交期","客户","工序"],
    "damo": ["是否核对","序列号","生产车间","工单编号","工单状态","订单批号","确定交期","订单数量","料品编码","生产线编号","规格型号","报工人","工单数量","报工数量","报废数量","返修数量","车间提供","时间","开工时间","完工时间","客户","工序"],
    "damo_dm": ["是否核对","序列号","生产车间","工单编号","工单状态","订单批号","确定交期","订单数量","料品编码","料品名称","生产线编号","规格型号","报工人","工单数量","报工数量","报废数量","返修数量","车间提供","时间","开工时间","完工时间","客户","工序","急货"],
}

def _export_excel(data, columns, filename):
    from urllib.parse import quote
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            val = row_data.get(col_name, "")
            if isinstance(val, str) and len(val) > 100:
                val = val[:100]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_name in ["订单数量","工单数量","报工数量","报废数量","返修数量","报工重量"]:
                cell.alignment = Alignment(horizontal='right')
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0'
            elif col_name in ["车间提供","时间"]:
                cell.alignment = Alignment(horizontal='right')
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0.0'
    for col_idx in range(1, len(columns) + 1):
        max_len = len(str(ws.cell(row=1, column=col_idx).value))
        for row in range(2, len(data) + 2):
            cell_val = ws.cell(row=row, column=col_idx).value
            if cell_val:
                max_len = max(max_len, min(len(str(cell_val)), 50))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 4
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )

def _export_query(sql, build_fn, start, end, 序列号=None, 姓名=None, 订单批号=None, columns_key=None, filename_prefix="报工详情"):
    extra = ""
    params = [start, end]
    if 序列号:
        extra += " AND b._Identify LIKE ?"
        params.append(f"%{序列号}%")
    if 姓名:
        extra += " AND a.emp_name LIKE ?"
        params.append(f"%{姓名}%")
    if 订单批号:
        extra += " AND a.OrderNumber LIKE ?"
        params.append(f"%{订单批号}%")
    data_sql = f"{sql}{extra} ORDER BY b.订单批号"
    rows = fetch_rows(data_sql, params)
    data = build_fn(rows)
    columns = EXPORT_COLUMNS.get(columns_key, list(data[0].keys()) if data else [])
    return _export_excel(data, columns, f"{filename_prefix}_{start}_{end}.xlsx")

@router.get("/workshopReportDetail/LockA/export", summary="锁体A报工详情导出")
async def export_lock_a(start: str = Query(...), end: str = Query(...), 序列号: Optional[str] = None, 姓名: Optional[str] = None, 订单批号: Optional[str] = None):
    return _export_query(SQL_A, build_a, start, end, 序列号, 姓名, 订单批号, "lock_a", "锁体A报工详情")

@router.get("/workshopReportDetail/LockB/export", summary="锁体B报工详情导出")
async def export_lock_b(start: str = Query(...), end: str = Query(...), 序列号: Optional[str] = None, 姓名: Optional[str] = None, 订单批号: Optional[str] = None):
    return _export_query(SQL_B, build_b, start, end, 序列号, 姓名, 订单批号, "lock_b", "锁体B报工详情")

@router.get("/workshopReportDetail/LockC/export", summary="锁体C报工详情导出")
async def export_lock_c(start: str = Query(...), end: str = Query(...), 序列号: Optional[str] = None, 姓名: Optional[str] = None, 订单批号: Optional[str] = None):
    return _export_query(SQL_C, build_c, start, end, 序列号, 姓名, 订单批号, "lock_c", "锁体C报工详情")

@router.get("/workshopReportDetail/LockD/export", summary="锁体D报工详情导出")
async def export_lock_d(start: str = Query(...), end: str = Query(...), 序列号: Optional[str] = None, 姓名: Optional[str] = None, 订单批号: Optional[str] = None):
    return _export_query(SQL_D, build_d, start, end, 序列号, 姓名, 订单批号, "lock_d", "锁体D报工详情")

@router.get("/workshopReportDetail/Suoliang/export", summary="锁梁报工详情导出")
async def export_suoliang(start: str = Query(...), end: str = Query(...), 序列号: Optional[str] = None, 姓名: Optional[str] = None, 订单批号: Optional[str] = None):
    return _export_query(SQL_SUOLIANG, build_suoliang, start, end, 序列号, 姓名, 订单批号, "suoliang", "锁梁报工详情")

@router.get("/workshopReportDetail/Material/export", summary="开料车间报工详情导出")
async def export_material(start: str = Query(...), end: str = Query(...), 工单编号: Optional[str] = None, 序列号: Optional[str] = None, 姓名: Optional[str] = None, 生产线编号: Optional[str] = None, 生产线描述: Optional[str] = None):
    extra = ""
    count_extra = ""
    params = [start, end]
    if 序列号:
        extra += " AND b._Identify LIKE ?"
        count_extra += " AND b._Identify LIKE ?"
        params.append(f"%{序列号}%")
    if 姓名:
        extra += " AND a.emp_name LIKE ?"
        count_extra += " AND a.emp_name LIKE ?"
        params.append(f"%{姓名}%")
    if 工单编号:
        extra += " AND a.JobExternalId LIKE ?"
        count_extra += " AND a.JobExternalId LIKE ?"
        params.append(f"%{工单编号}%")
    if 生产线编号:
        extra += " AND a.ResName LIKE ?"
        count_extra += " AND a.ResName LIKE ?"
        params.append(f"%{生产线编号}%")
    if 生产线描述:
        extra += " AND rc.Description LIKE ?"
        count_extra += " AND rc.Description LIKE ?"
        params.append(f"%{生产线描述}%")
    data_sql = f"{SQL_MATERIAL}{extra} ORDER BY b.订单批号"
    rows = fetch_rows(data_sql, params)
    data = build_material(rows)
    columns = EXPORT_COLUMNS.get("material", list(data[0].keys()) if data else [])
    return _export_excel(data, columns, f"开料车间报工详情_{start}_{end}.xlsx")

@router.get("/workshopReportDetail/Zhuangqian/export", summary="装嵌报工详情导出")
async def export_zhuangqian(start: str = Query(...), end: str = Query(...), 序列号: Optional[str] = None, 姓名: Optional[str] = None, 工单批号: Optional[str] = None, 锁类分区: Optional[str] = None, 含TSA: bool = False):
    extra = ""
    params = [start, end]
    if 序列号:
        extra += " AND b._Identify LIKE ?"
        params.append(f"%{序列号}%")
    if 姓名:
        extra += " AND a.emp_name LIKE ?"
        params.append(f"%{姓名}%")
    if 工单批号:
        extra += " AND a.OrderNumber LIKE ?"
        params.append(f"%{工单批号}%")
    if 锁类分区:
        extra += " AND b.锁类分区 LIKE ?"
        params.append(f"%{锁类分区}%")
    if not 含TSA:
        extra += " AND b.料品名称 <> 'TSA006'"
    data_sql = f"{SQL_ZHUANGQIAN}{extra} ORDER BY b.订单批号"
    rows = fetch_rows(data_sql, params)
    data = build_zhuangqian(rows)
    columns = EXPORT_COLUMNS.get("zhuangqian", list(data[0].keys()) if data else [])
    return _export_excel(data, columns, f"装嵌报工详情_{start}_{end}.xlsx")

@router.get("/workshopReportDetail/Damo/export", summary="打磨-装配区报工详情导出")
async def export_damo(start: str = Query(...), end: str = Query(...), 序列号: Optional[str] = None, 姓名: Optional[str] = None, 生产线编号: Optional[str] = None, 工单批号: Optional[str] = None):
    extra = ""
    params = [start, end]
    if 序列号:
        extra += " AND b._Identify LIKE ?"
        params.append(f"%{序列号}%")
    if 姓名:
        extra += " AND a.emp_name LIKE ?"
        params.append(f"%{姓名}%")
    if 生产线编号:
        extra += " AND a.ResName LIKE ?"
        params.append(f"%{生产线编号}%")
    if 工单批号:
        extra += " AND a.OrderNumber LIKE ?"
        params.append(f"%{工单批号}%")
    data_sql = f"{SQL_DAMO}{extra} ORDER BY b.订单批号"
    rows = fetch_rows(data_sql, params)
    data = build_damo(rows)
    columns = EXPORT_COLUMNS.get("damo", list(data[0].keys()) if data else [])
    return _export_excel(data, columns, f"打磨装配区报工详情_{start}_{end}.xlsx")

@router.get("/workshopReportDetail/DamoDM/export", summary="打磨-打磨区报工详情导出")
async def export_damo_dm(start: str = Query(...), end: str = Query(...), 序列号: Optional[str] = None, 姓名: Optional[str] = None, 规格型号: Optional[str] = None, 工序: Optional[str] = None, 生产线编号: Optional[str] = None, 工单批号: Optional[str] = None):
    extra = ""
    params = [start, end]
    if 序列号:
        extra += " AND b._Identify LIKE ?"
        params.append(f"%{序列号}%")
    if 姓名:
        extra += " AND a.emp_name LIKE ?"
        params.append(f"%{姓名}%")
    if 规格型号:
        extra += " AND a.ProductDescription LIKE ?"
        params.append(f"%{规格型号}%")
    if 工序:
        extra += " AND b.OpExternalId LIKE ?"
        params.append(f"%{工序}%")
    if 生产线编号:
        extra += " AND a.ResName LIKE ?"
        params.append(f"%{生产线编号}%")
    if 工单批号:
        extra += " AND a.OrderNumber LIKE ?"
        params.append(f"%{工单批号}%")
    data_sql = f"{SQL_DAMO_DM}{extra} ORDER BY b.订单批号"
    rows = fetch_rows(data_sql, params)
    data = build_damo_dm(rows)
    columns = EXPORT_COLUMNS.get("damo_dm", list(data[0].keys()) if data else [])
    return _export_excel(data, columns, f"打磨区报工详情_{start}_{end}.xlsx")
