from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from typing import Optional
import pyodbc
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io

router = APIRouter()

DB_SERVER = "192.168.41.57"
DB_DATABASE = "department2020"
DB_USERNAME = "sa"
DB_PASSWORD = "3518i"

def get_db_connection():
    conn_str = (
        f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={DB_SERVER};DATABASE={DB_DATABASE};"
        f"UID={DB_USERNAME};PWD={DB_PASSWORD};"
    )
    return pyodbc.connect(conn_str)

def get_data_from_db(start, end, 序列号=None, 姓名=None, 生产线编号=None, 订单批号=None, page=1, page_size=100):
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        base_sql = """
        SELECT a.isCheck, a.iNo, b._Identify, b.生产车间, a.JobExternalId AS 工单编号,
               b.工单状态, a.OrderNumber AS 订单批号, b.订单数量,
               a.ItemExternalId AS 料品编码, a.ResName AS 生产线编号,
               a.ProductDescription AS 规格型号, a.emp_name AS 报工人,
               b.计划产量 AS 工单数量, a.EachFinishedQty AS 报工数量,
               a.scrapQty AS 报废数量, a.repairQty AS 返修数量,
               CASE WHEN d.capacity IS NOT NULL THEN d.capacity ELSE 0 END AS 每小时产能,
               a.StartDate AS 开工时间, a.FinishedDate AS 完工时间,
               b.确定交期, b.客户, b.OpExternalId
        FROM APS_FinishedQty_Pack a
        JOIN 派工单 b ON a.JobExternalId = b.工单编号
        LEFT JOIN APS.APS_SUO.dbo.offline_process d
            ON b.OpExternalId = d.proccess
            AND b.料品编码 = d.item_no + d.proccess
        WHERE a.EachFinishedQty > 0
        AND a.FinishedDate BETWEEN ? AND ?
        """
        extra = ""
        params = [start, end]
        if 序列号:
            extra += " AND b._Identify LIKE ?"
            params.append(f"%{序列号}%")
        if 姓名:
            extra += " AND a.emp_name LIKE ?"
            params.append(f"%{姓名}%")
        if 生产线编号:
            extra += " AND a.ResName LIKE ?"
            params.append(f"%{生产线编号}%")
        if 订单批号:
            extra += " AND a.OrderNumber LIKE ?"
            params.append(f"%{订单批号}%")
        count_sql = "SELECT COUNT(*) FROM APS_FinishedQty_Pack a JOIN 派工单 b ON a.JobExternalId = b.工单编号 WHERE a.EachFinishedQty > 0 AND a.FinishedDate BETWEEN ? AND ?"
        extra_count = ""
        count_params = list(params)
        if 序列号:
            extra_count += " AND b._Identify LIKE ?"
            count_params.append(f"%{序列号}%")
        if 姓名:
            extra_count += " AND a.emp_name LIKE ?"
            count_params.append(f"%{姓名}%")
        if 生产线编号:
            extra_count += " AND a.ResName LIKE ?"
            count_params.append(f"%{生产线编号}%")
        if 订单批号:
            extra_count += " AND a.OrderNumber LIKE ?"
            count_params.append(f"%{订单批号}%")
        cursor.execute(count_sql + extra_count, count_params)
        total_count = cursor.fetchone()[0]
        offset = (page - 1) * page_size
        data_sql = f"{base_sql}{extra} ORDER BY b.订单批号 OFFSET ? ROWS FETCH NEXT ? ROWS ONLY"
        data_params = list(params) + [offset, page_size]
        cursor.execute(data_sql, data_params)
        rows = cursor.fetchall()

        result = []
        for row in rows:
            报工数量 = float(row[13]) if row[13] else 0
            每小时产能 = float(row[16]) if row[16] else 0
            时间 = round(报工数量 / 每小时产能, 2) if 每小时产能 != 0 else None

            result.append({
                "是否核对": row[0] or "",
                "序列号": row[2] or "",
                "生产车间": row[3] or "",
                "工单编号": row[4] or "",
                "工单状态": row[5] or "",
                "订单批号": row[6] or "",
                "订单数量": float(row[7]) if row[7] else 0,
                "料品编码": row[8] or "",
                "生产线编号": row[9] or "",
                "规格型号": row[10] or "",
                "报工人": row[11] or "",
                "工单数量": float(row[12]) if row[12] else 0,
                "报工数量": 报工数量,
                "报废数量": float(row[14]) if row[14] else 0,
                "返修数量": float(row[15]) if row[15] else 0,
                "车间提供": 每小时产能,
                "时间": 时间,
                "开工时间": str(row[17]) if row[17] else "",
                "完工时间": str(row[18]) if row[18] else "",
                "确定交期": row[19] or "",
                "客户": row[20] or "",
                "工序": row[21] or "",
            })
        return result, total_count
    finally:
        if conn:
            try: conn.close()
            except: pass

@router.get("/workshopReportDetail/Pack", summary="包装车间报工详情")
async def get_data(
    start: str = Query(..., description="开始时间"),
    end: str = Query(..., description="结束时间"),
    序列号: Optional[str] = Query(None),
    姓名: Optional[str] = Query(None),
    生产线编号: Optional[str] = Query(None),
    订单批号: Optional[str] = Query(None),
    page: int = 1,
    page_size: int = 100,
):
    try:
        raw_data, total_count = get_data_from_db(start, end, 序列号, 姓名, 生产线编号, 订单批号, page, page_size)
        return {"status": "success", "data": raw_data, "total_count": total_count}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"服务器错误: {exc}")

PACK_COLUMNS = ["是否核对","序列号","生产车间","工单编号","工单状态","订单批号","订单数量","料品编码","生产线编号","规格型号","报工人","工单数量","报工数量","报废数量","返修数量","车间提供","时间","开工时间","完工时间","确定交期","客户","工序"]

def _export_excel(data, columns, filename):
    from urllib.parse import quote
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            val = row_data.get(col_name, "")
            if isinstance(val, str) and len(val) > 100:
                val = val[:100]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_name in ["订单数量","工单数量","报工数量","报废数量","返修数量"]:
                cell.alignment = Alignment(horizontal='right')
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0'
            elif col_name in ["车间提供","时间"]:
                cell.alignment = Alignment(horizontal='right')
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0.0'
    for col_idx in range(1, len(columns) + 1):
        max_len = len(str(ws.cell(row=1, column=col_idx).value))
        for row in range(2, len(data) + 2):
            cell_val = ws.cell(row=row, column=col_idx).value
            if cell_val:
                max_len = max(max_len, min(len(str(cell_val)), 50))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 4
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )

@router.get("/workshopReportDetail/Pack/export", summary="包装车间报工详情导出")
async def export_pack(
    start: str = Query(...), end: str = Query(...),
    序列号: Optional[str] = Query(None), 姓名: Optional[str] = Query(None),
    生产线编号: Optional[str] = Query(None), 订单批号: Optional[str] = Query(None),
):
    try:
        base_sql = """
        SELECT a.isCheck, a.iNo, b._Identify, b.生产车间, a.JobExternalId AS 工单编号,
               b.工单状态, a.OrderNumber AS 订单批号, b.订单数量,
               a.ItemExternalId AS 料品编码, a.ResName AS 生产线编号,
               a.ProductDescription AS 规格型号, a.emp_name AS 报工人,
               b.计划产量 AS 工单数量, a.EachFinishedQty AS 报工数量,
               a.scrapQty AS 报废数量, a.repairQty AS 返修数量,
               CASE WHEN d.capacity IS NOT NULL THEN d.capacity ELSE 0 END AS 每小时产能,
               a.StartDate AS 开工时间, a.FinishedDate AS 完工时间,
               b.确定交期, b.客户, b.OpExternalId
        FROM APS_FinishedQty_Pack a
        JOIN 派工单 b ON a.JobExternalId = b.工单编号
        LEFT JOIN APS.APS_SUO.dbo.offline_process d
            ON b.OpExternalId = d.proccess AND b.料品编码 = d.item_no + d.proccess
        WHERE a.EachFinishedQty > 0 AND a.FinishedDate BETWEEN ? AND ?
        """
        extra = ""
        params = [start, end]
        if 序列号:
            extra += " AND b._Identify LIKE ?"
            params.append(f"%{序列号}%")
        if 姓名:
            extra += " AND a.emp_name LIKE ?"
            params.append(f"%{姓名}%")
        if 生产线编号:
            extra += " AND a.ResName LIKE ?"
            params.append(f"%{生产线编号}%")
        if 订单批号:
            extra += " AND a.OrderNumber LIKE ?"
            params.append(f"%{订单批号}%")
        data_sql = f"{base_sql}{extra} ORDER BY b.订单批号"
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute(data_sql, params)
            rows = cursor.fetchall()
            data = []
            for row in rows:
                报工数量 = float(row[13]) if row[13] else 0
                每小时产能 = float(row[16]) if row[16] else 0
                时间 = round(报工数量 / 每小时产能, 2) if 每小时产能 != 0 else None
                data.append({
                    "是否核对": row[0] or "", "序列号": row[2] or "", "生产车间": row[3] or "",
                    "工单编号": row[4] or "", "工单状态": row[5] or "", "订单批号": row[6] or "",
                    "订单数量": float(row[7]) if row[7] else 0, "料品编码": row[8] or "",
                    "生产线编号": row[9] or "", "规格型号": row[10] or "", "报工人": row[11] or "",
                    "工单数量": float(row[12]) if row[12] else 0, "报工数量": 报工数量,
                    "报废数量": float(row[14]) if row[14] else 0, "返修数量": float(row[15]) if row[15] else 0,
                    "车间提供": 每小时产能, "时间": 时间,
                    "开工时间": str(row[17]) if row[17] else "", "完工时间": str(row[18]) if row[18] else "",
                    "确定交期": row[19] or "", "客户": row[20] or "", "工序": row[21] or "",
                })
        finally:
            conn.close()
        return _export_excel(data, PACK_COLUMNS, f"包装车间报工详情_{start}_{end}.xlsx")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"服务器错误: {exc}")
