from fastapi import APIRouter, HTTPException, Query
from typing import Optional
import pyodbc
from datetime import datetime
from fastapi_cache.decorator import cache
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from collections import defaultdict
import io
import urllib.parse

router = APIRouter()

DB_SERVER = "192.168.10.200"
DB_USERNAME = "sa"
DB_PASSWORD = "5tgb^YHN7ujm*IK<"

def get_db_connection():
    conn_str = (
        f"DRIVER={{SQL Server}};SERVER={DB_SERVER};DATABASE=APS_SUO;"
        f"UID={DB_USERNAME};PWD={DB_PASSWORD};"
    )
    return pyodbc.connect(conn_str)

COLUMNS = [
    "物料类型", "料品编码", "料品名称", "料品规格",
    "生产车间", "工序", "工序规格码", "生效单价"
]

COLUMN_MAP = {
    "物料类型": "a.[data_type]",
    "料品编码": "a.[item_no]",
    "料品名称": "a.[part_name]",
    "料品规格": "a.[part_spec]",
    "生产车间": "a.[生产车间]",
    "工序": "a.[proccess]",
    "工序规格码": "a.[proccessNumber]",
    "生效单价": "b.[生效单价]"
}

def get_data_from_db(part_name=None, 生产车间=None, 工序=None):
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        select_cols = ", ".join(COLUMN_MAP.values())
        sql = f"""
        SELECT {select_cols}
        FROM [APS_SUO].[dbo].[offline_process] a
        LEFT JOIN [TSD].[wygz].[dbo].[工序单价表] b
        ON a.proccessNumber = b.工序规格码
        WHERE a.proccessNumber IS NOT NULL AND b.生效单价 IS NOT NULL
        """
        params = []
        if part_name:
            sql += " AND a.[part_name] LIKE ?"
            params.append(f"%{part_name}%")
        if 生产车间:
            sql += " AND LTRIM(RTRIM(a.[生产车间])) = ?"
            params.append(生产车间.strip())
        if 工序:
            sql += " AND a.[proccess] LIKE ?"
            params.append(f"%{工序}%")

        sql += " ORDER BY a.[data_type], a.[item_no]"

        cursor.execute(sql, params)
        rows = cursor.fetchall()

        raw_data = []
        for row in rows:
            raw_data.append({
                "物料类型": row[0] if row[0] else "",
                "料品编码": row[1] if row[1] else "",
                "料品名称": row[2] if row[2] else "",
                "料品规格": row[3] if row[3] else "",
                "生产车间": row[4] if row[4] else "",
                "工序": row[5] if row[5] else "",
                "工序规格码": row[6] if row[6] else "",
                "生效单价": float(row[7]) if row[7] else 0,
            })

        return raw_data
    except Exception as exc:
        raise exc
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass

def get_cascade_options(part_name=None, 生产车间=None, 工序=None):
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        where_parts = ["a.proccessNumber IS NOT NULL", "EXISTS (SELECT 1 FROM [TSD].[wygz].[dbo].[工序单价表] b WHERE a.proccessNumber = b.工序规格码 AND b.生效单价 IS NOT NULL)"]
        params = []

        if part_name:
            where_parts.append("a.[part_name] LIKE ?")
            params.append(f"%{part_name}%")
        if 生产车间:
            where_parts.append("LTRIM(RTRIM(a.[生产车间])) = ?")
            params.append(生产车间.strip())
        if 工序:
            where_parts.append("a.[proccess] LIKE ?")
            params.append(f"%{工序}%")

        where_clause = " AND ".join(where_parts)

        # 料品名称选项
        sql_name = f"SELECT DISTINCT a.[part_name] AS v FROM [APS_SUO].[dbo].[offline_process] a WHERE {where_clause} AND a.[part_name] IS NOT NULL AND a.[part_name] != '' ORDER BY v"
        cursor.execute(sql_name, params)
        料品名称_options = [row[0] for row in cursor.fetchall()]

        # 生产车间选项
        sql_dept = f"SELECT DISTINCT TOP 200 LTRIM(RTRIM(a.[生产车间])) AS v FROM [APS_SUO].[dbo].[offline_process] a WHERE {where_clause} AND a.[生产车间] IS NOT NULL AND LTRIM(RTRIM(a.[生产车间])) != '' ORDER BY v"
        cursor.execute(sql_dept, params)
        生产车间_options = [row[0] for row in cursor.fetchall()]

        # 工序选项
        sql_proc = f"SELECT DISTINCT TOP 200 LTRIM(RTRIM(a.[proccess])) AS v FROM [APS_SUO].[dbo].[offline_process] a WHERE {where_clause} AND a.[proccess] IS NOT NULL AND LTRIM(RTRIM(a.[proccess])) != '' ORDER BY v"
        cursor.execute(sql_proc, params)
        工序_options = [row[0] for row in cursor.fetchall()]

        return {
            "生产车间": 生产车间_options,
            "工序": 工序_options,
            "料品名称": 料品名称_options,
        }
    except Exception as exc:
        raise exc
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass

def compute_stats(data):
    grouped = defaultdict(lambda: {"规格": set(), "工序": set()})
    for item in data:
        ws = item.get("生产车间", "")
        spec = item.get("料品规格", "")
        proc = item.get("工序", "")
        if ws:
            if spec:
                grouped[ws]["规格"].add(spec)
            if proc:
                grouped[ws]["工序"].add(proc)

    stats = []
    for ws in sorted(grouped.keys()):
        specs = sorted(grouped[ws]["规格"])
        procs = sorted(grouped[ws]["工序"])
        stats.append({
            "生产车间": ws,
            "规格数": len(specs),
            "工序数": len(procs),
            "规格列表": specs,
            "工序列表": procs,
        })
    return stats

@router.get("/offlineProcess", summary="排产所有工序对应单价")
@cache(expire=1800)
async def get_offline_process(
    料品名称: Optional[str] = Query(None, description="料品名称（模糊查询）"),
    生产车间: Optional[str] = Query(None, description="车间（精确匹配）"),
    工序: Optional[str] = Query(None, description="工序（模糊查询）"),
):
    try:
        raw_data = get_data_from_db(料品名称, 生产车间, 工序)
        stats = compute_stats(raw_data)
        return {
            "status": "success",
            "data": raw_data,
            "stats": stats,
            "total_count": len(raw_data),
            "timestamp": datetime.now().isoformat()
        }
    except Exception as exc:
        print(f"获取数据失败: {exc}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {exc}")

@router.get("/offlineProcess/cascade-options", summary="级联筛选下拉选项")
@cache(expire=1800)
async def get_cascade_options_endpoint(
    料品名称: Optional[str] = Query(None, description="料品名称"),
    生产车间: Optional[str] = Query(None, description="车间"),
    工序: Optional[str] = Query(None, description="工序"),
):
    try:
        data = get_cascade_options(料品名称, 生产车间, 工序)
        return {"status": "success", "data": data}
    except Exception as exc:
        print(f"获取级联选项失败: {exc}")
        return {"status": "success", "data": {"生产车间": [], "工序": [], "料品名称": []}}

@router.get("/offlineProcess/export", summary="导出Excel")
async def export_offline_process_excel(
    料品名称: Optional[str] = Query(None, description="料品名称"),
    生产车间: Optional[str] = Query(None, description="车间"),
    工序: Optional[str] = Query(None, description="工序"),
):
    try:
        data = get_data_from_db(料品名称, 生产车间, 工序)

        wb = Workbook()
        ws = wb.active
        ws.title = "排产所有工序对应单价"

        headers = COLUMNS
        ws.append(headers)

        for row in data:
            ws.append([row.get(c, "") for c in COLUMNS])

        for idx, col_name in enumerate(headers, start=1):
            col_letter = get_column_letter(idx)
            max_len = len(str(col_name))
            for row in data:
                cell_val = str(row.get(col_name, ""))
                if len(cell_val) > max_len:
                    max_len = len(cell_val)
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = "排产所有工序对应单价.xlsx"
        encoded_filename = urllib.parse.quote(filename)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"导出失败: {exc}")

@router.get("/offlineProcess/options", summary="下拉选项（兼容旧版）")
async def get_options(field: str = Query(..., description="字段名")):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        field_map = {
            "data_type": ("a.[data_type]", "a.[data_type]"),
            "part_name": ("a.[part_name]", "a.[part_name]"),
            "DepartmentName": ("LTRIM(RTRIM(a.[DepartmentName]))", "a.[DepartmentName]"),
            "proccess": ("a.[proccess]", "a.[proccess]"),
        }

        if field not in field_map:
            return {"status": "success", "data": []}

        select_expr, filter_expr = field_map[field]
        sql = f"""
        SELECT DISTINCT TOP 200 {select_expr} AS value
        FROM [APS_SUO].[dbo].[offline_process] a
        LEFT JOIN [TSD].[wygz].[dbo].[工序单价表] b
        ON a.proccessNumber = b.工序规格码
        WHERE a.proccessNumber IS NOT NULL AND b.生效单价 IS NOT NULL
          AND {filter_expr} IS NOT NULL AND {filter_expr} != ''
        ORDER BY {select_expr}
        """
        cursor.execute(sql)
        result = [{"value": row[0]} for row in cursor.fetchall()]
        conn.close()
        return {"status": "success", "data": result}
    except Exception as exc:
        print(f"获取选项失败: field={field}, error={exc}")
        return {"status": "success", "data": []}
