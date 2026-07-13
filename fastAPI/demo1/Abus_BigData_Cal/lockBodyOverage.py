import logging
import time
from fastapi import APIRouter, HTTPException, Query
from typing import Optional
import pyodbc
from datetime import datetime
from collections import defaultdict
import re
import io
import urllib.parse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from fastapi.responses import StreamingResponse

logger = logging.getLogger(__name__)

router = APIRouter()

# 简单内存缓存：每5分钟刷新一次
_cached_data = None
_cached_time = 0
CACHE_TTL = 300  # 5分钟

def get_cached_data():
    global _cached_data, _cached_time
    now = time.time()
    if _cached_data is not None and (now - _cached_time) < CACHE_TTL:
        logger.info(f"使用缓存数据，距离上次更新 {now - _cached_time:.0f}秒")
        return _cached_data
    _cached_data = get_all_data()
    _cached_time = time.time()
    logger.info(f"缓存已刷新，共 {len(_cached_data)} 条数据")
    return _cached_data

ERP_SERVER = "192.168.1.1"
ERP_DATABASE = "huayueerp"
ERP_USER = "sa"
ERP_PASSWORD = "3518i"

def get_erp_connection():
    conn_str = (
        f"DRIVER={{SQL Server}};SERVER={ERP_SERVER};DATABASE={ERP_DATABASE};"
        f"UID={ERP_USER};PWD={ERP_PASSWORD};"
    )
    return pyodbc.connect(conn_str)

ORDER_RE = re.compile(r"\u3011([^|]+)\|")

def parse_order_no(rem):
    if rem is None:
        return None
    text = str(rem).strip()
    match = ORDER_RE.search(text)
    if match:
        return match.group(1).strip()
    if "|" in text:
        return text.split("|", 1)[0].split("\u3011")[-1].strip() or None
    return None

COLUMNS = [
    "订单批号", "锁类分区", "锁体品号", "锁体物料编号", "锁体物料名称", "锁体规格",
    "订单需求数量", "锁体历史入库数量", "超入库数量", "是否超入库", "入库完成率",
    "订单单位", "入库单位", "入库仓库", "最早入库日期", "最后入库日期",
    "要求交期", "确定交期", "订单状态",
]

def get_all_data():
    conn = None
    try:
        logger.info("开始查询ERP数据库...")
        conn = get_erp_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT i.part_no, p.item_no, p.part_name, p.part_spec,
                   wh.wh_name, i.sheet_qty, i.unit_no, i.pro_date, i.rem
            FROM oscm_other2 i WITH (NOLOCK)
            INNER JOIN obas_part p WITH (NOLOCK) ON i.part_no = p.part_no
            LEFT JOIN obas_wh wh WITH (NOLOCK) ON i.wh_no = wh.wh_no
            WHERE p.part_name LIKE ?
              AND LEFT(CONVERT(varchar(50), p.item_no), 1) = '2'
              AND i.rem IS NOT NULL
        """, ("%锁体%",))
        inbound_rows = cursor.fetchall()

        order_nos = set()
        inbound_by_order = defaultdict(list)
        for row in inbound_rows:
            order_no = parse_order_no(row[8])
            if not order_no:
                continue
            order_no = str(order_no)[:100]
            order_nos.add(order_no)
            inbound_by_order[order_no].append({
                "part_no": row[0], "item_no": row[1],
                "part_name": row[2], "part_spec": row[3],
                "wh_name": row[4], "sheet_qty": float(row[5] or 0),
                "unit_no": row[6], "pro_date": row[7],
            })

        if not order_nos:
            return []

        cursor.execute("CREATE TABLE #order_nos (order_no nvarchar(100) NOT NULL PRIMARY KEY)")
        cursor.fast_executemany = True
        cursor.executemany("INSERT INTO #order_nos (order_no) VALUES (?)", [(v,) for v in order_nos])
        cursor.fast_executemany = False

        cursor.execute("""
            SELECT o.[订单批号], o.[锁类分区],
                   o.[订单数量], o.[销售单位],
                   o.[要求交期], o.[确定交期], o.[订单状态]
            FROM [dbo].[订单-订单明细] o WITH (NOLOCK)
            INNER JOIN #order_nos n ON o.[订单批号] = n.order_no
            WHERE o.[锁类分区] = ?
        """, ("铝门锁",))
        order_rows = cursor.fetchall()
        cursor.execute("DROP TABLE #order_nos")

        orders = {}
        for row in order_rows:
            on = row[0]
            if on not in orders:
                orders[on] = {
                    "lock_category": row[1], "order_qty": float(row[2] or 0),
                    "sales_unit": row[3], "request_date": str(row[4] or ""),
                    "confirm_delivery_date": str(row[5] or ""), "order_status": row[6],
                }

        stats = defaultdict(lambda: {
            "item_no": "", "part_name": "", "part_spec": "",
            "inbound_qty": 0, "first_in_date": None, "last_in_date": None,
            "warehouses": set(), "inbound_unit": "",
        })
        for order_no, items in inbound_by_order.items():
            if order_no not in orders:
                continue
            for item in items:
                key = (order_no, item["part_no"])
                s = stats[key]
                if item["item_no"]:
                    s["item_no"] = item["item_no"]
                if item["part_name"]:
                    s["part_name"] = item["part_name"]
                if item["part_spec"]:
                    s["part_spec"] = item["part_spec"]
                s["inbound_qty"] += item["sheet_qty"]
                if s["first_in_date"] is None or item["pro_date"] < s["first_in_date"]:
                    s["first_in_date"] = item["pro_date"]
                if s["last_in_date"] is None or item["pro_date"] > s["last_in_date"]:
                    s["last_in_date"] = item["pro_date"]
                s["warehouses"].add(item["wh_name"])
                s["inbound_unit"] = item["unit_no"]

        result = []
        for (order_no, part_no), s in stats.items():
            o = orders[order_no]
            overage_qty = s["inbound_qty"] - o["order_qty"]
            if overage_qty <= 0:
                continue
            result.append({
                "订单批号": order_no,
                "锁类分区": o["lock_category"],
                "锁体品号": s["item_no"],
                "锁体物料编号": part_no,
                "锁体物料名称": s["part_name"],
                "锁体规格": s["part_spec"] or "",
                "订单需求数量": o["order_qty"],
                "锁体历史入库数量": s["inbound_qty"],
                "超入库数量": overage_qty,
                "是否超入库": "是",
                "入库完成率": round(s["inbound_qty"] / o["order_qty"], 6) if o["order_qty"] else None,
                "订单单位": o["sales_unit"],
                "入库单位": s["inbound_unit"],
                "入库仓库": "、".join(sorted([w for w in s["warehouses"] if w])),
                "最早入库日期": str(s["first_in_date"] or ""),
                "最后入库日期": str(s["last_in_date"] or ""),
                "要求交期": o["request_date"],
                "确定交期": o["confirm_delivery_date"],
                "订单状态": o["order_status"],
            })

        result.sort(key=lambda x: (x["订单批号"], x["锁体品号"], x["锁体物料编号"]))
        return result
    except Exception as exc:
        logger.error(f"get_all_data失败: {exc}", exc_info=True)
        raise exc
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass

def get_options_from_cache(field):
    data = get_cached_data()
    values = set()
    for row in data:
        v = row.get(field, "")
        if v:
            values.add(v)
    return sorted(values)

@router.get("/lockBodyOverage", summary="超市多出的损耗数")
async def get_lock_body_overage(
    订单批号: Optional[str] = Query(None, description="订单批号"),
    物料编码: Optional[str] = Query(None, description="物料编码（按锁体品号过滤）"),
    锁体规格: Optional[str] = Query(None, description="锁体规格"),
):
    try:
        logger.info(f"收到查询请求: 订单批号={订单批号}, 物料编码={物料编码}, 锁体规格={锁体规格}")
        all_data = get_cached_data()
        logger.info(f"查询到 {len(all_data)} 条原始数据")
        if not all_data:
            return {"status": "success", "data": [], "total_count": 0, "timestamp": datetime.now().isoformat()}

        result = all_data
        if 订单批号:
            result = [r for r in result if r["订单批号"] == 订单批号]
        if 物料编码:
            result = [r for r in result if r["锁体品号"] == 物料编码]
        if 锁体规格:
            result = [r for r in result if r["锁体规格"] == 锁体规格]

        logger.info(f"返回 {len(result)} 条数据")
        return {
            "status": "success",
            "data": result,
            "total_count": len(result),
            "timestamp": datetime.now().isoformat()
        }
    except Exception as exc:
        logger.error(f"get_lock_body_overage失败: {exc}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"服务器错误: {exc}")

@router.get("/lockBodyOverage/options", summary="下拉选项")
async def get_options(field: str = Query(..., description="字段名")):
    try:
        valid = ["订单批号", "锁体品号", "锁体规格"]
        if field not in valid:
            return {"status": "success", "data": []}
        logger.info(f"查询options: field={field}")
        data = [{"value": v} for v in get_options_from_cache(field)]
        logger.info(f"options {field} 返回 {len(data)} 条")
        return {"status": "success", "data": data}
    except Exception as exc:
        logger.error(f"get_options失败 field={field}: {exc}", exc_info=True)
        return {"status": "success", "data": []}

@router.get("/lockBodyOverage/export", summary="导出Excel")
async def export_lock_body_overage(
    订单批号: Optional[str] = Query(None, description="订单批号"),
    物料编码: Optional[str] = Query(None, description="物料编码（按锁体品号过滤）"),
    锁体规格: Optional[str] = Query(None, description="锁体规格"),
):
    try:
        logger.info("开始导出Excel")
        all_data = get_cached_data()
        data = all_data
        if 订单批号:
            data = [r for r in data if r["订单批号"] == 订单批号]
        if 物料编码:
            data = [r for r in data if r["锁体品号"] == 物料编码]
        if 锁体规格:
            data = [r for r in data if r["锁体规格"] == 锁体规格]

        wb = Workbook()
        ws = wb.active
        ws.title = "超市多出的损耗数"
        ws.append(COLUMNS)
        for row in data:
            ws.append([row.get(c, "") for c in COLUMNS])

        for idx in range(1, len(COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 20

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = "超市多出的损耗数.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{urllib.parse.quote(filename)}"}
        )
    except Exception as exc:
        logger.error(f"导出失败: {exc}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导出失败: {exc}")
