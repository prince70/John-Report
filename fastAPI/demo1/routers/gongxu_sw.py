from datetime import datetime
import os
from typing import Optional
import io
import urllib.parse
import re
import json

import pyodbc
from fastapi import APIRouter, HTTPException, status, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

router = APIRouter()
DB_SERVER = "192.168.10.200"
DB_DATABASE = "APS_SUO"
DB_USERNAME = "sa"
DB_PASSWORD = "5tgb^YHN7ujm*IK<"


def get_db_connection():
    server = DB_SERVER
    database = DB_DATABASE
    username = DB_USERNAME
    password = DB_PASSWORD

    trusted_conn_str = (
        f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};"
        f"DATABASE={database};Trusted_Connection=yes;"
    )
    sql_auth_conn_str = (
        f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};"
        f"UID={username};PWD={password};"
    )

    sql_auth_error = None
    trusted_error = None
    use_trusted = False

    def try_sql_auth():
        nonlocal sql_auth_error
        try:
            return pyodbc.connect(sql_auth_conn_str)
        except Exception as exc:
            sql_auth_error = exc
            return None

    def try_trusted_auth():
        nonlocal trusted_error
        try:
            return pyodbc.connect(trusted_conn_str)
        except Exception as exc:
            trusted_error = exc
            return None

    # 默认优先 SQL 账号认证；仅当显式设置时优先集成认证。
    if use_trusted:
        conn = try_trusted_auth()
        if conn:
            return conn
        conn = try_sql_auth()
        if conn:
            return conn
    else:
        conn = try_sql_auth()
        if conn:
            return conn
        conn = try_trusted_auth()
        if conn:
            return conn

    raise HTTPException(
        status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        detail=(
            "[report-router-v3] "
            f"数据库连接失败。SQL认证错误: {sql_auth_error}; "
            f"集成认证错误: {trusted_error}. "
            "请检查 DB_SERVER_REPORT/DB_DATABASE_REPORT/DB_USERNAME_REPORT/DB_PASSWORD_REPORT"
        ),
    )


def parse_csv_values(raw: Optional[str]):
    if not raw:
        return []
    text = str(raw).strip()

    # 支持 JSON 数组格式，允许将包含逗号的车间值作为一个整体传入。
    if text.startswith("[") and text.endswith("]"):
        try:
            parsed = json.loads(text)
            if isinstance(parsed, list):
                return [str(item).strip() for item in parsed if str(item).strip()]
        except Exception:
            pass

    text = text.replace("，", ",")
    return [item.strip() for item in text.split(",") if item and item.strip()]


def get_workshop_column(db_columns):
    return next((col for col in db_columns if "车间" in str(col)), None)


def get_process_name_columns(db_columns):
    columns = []
    for col in db_columns:
        name = str(col)
        is_legacy = re.search(r"第\s*\d+\s*道工序", name)
        is_renamed = ("开料尾工序" in name) or ("锁体A首工序" in name)
        if (is_legacy or is_renamed) and not re.search(r"报工数|结束时间|最早结束时间|最晚结束时间|时间|日期", name):
            columns.append(col)
    return columns


def build_report_where(
    item_code: Optional[str],
    product_name: Optional[str],
    product_spec: Optional[str],
    db_columns,
    workshop_values: Optional[str] = None,
    process_values: Optional[str] = None,
    date_column: Optional[str] = None,
    date_start: Optional[str] = None,
    date_end: Optional[str] = None,
    column_filters: Optional[str] = None,
):
    conditions = []
    params = []

    if item_code:
        conditions.append("[料品编码] LIKE ?")
        params.append(f"%{item_code}%")
    if product_name:
        conditions.append("[产品名称] LIKE ?")
        params.append(f"%{product_name}%")
    if product_spec:
        conditions.append("[产品规格] LIKE ?")
        params.append(f"%{product_spec}%")

    selected_workshops = [str(item).replace(" ", "") for item in parse_csv_values(workshop_values) if str(item).strip()]
    workshop_column = get_workshop_column(db_columns)
    if selected_workshops and workshop_column:
        workshop_col_sql = escape_sql_identifier(workshop_column)
        normalized_workshop_expr = (
            f"REPLACE(REPLACE(CONVERT(NVARCHAR(255), {workshop_col_sql}), N'，', N','), N' ', N'')"
        )

        workshop_ors = []
        for workshop in selected_workshops:
            # 支持字段里为 "A,B" 这类拼接值，避免精确 IN 造成筛选无数据。
            workshop_ors.append(f"(',' + {normalized_workshop_expr} + ',') LIKE ?")
            params.append(f"%,{workshop},%")
        conditions.append("(" + " OR ".join(workshop_ors) + ")")

    selected_processes = parse_csv_values(process_values)
    process_columns = get_process_name_columns(db_columns)
    if selected_processes and process_columns:
        placeholders = ", ".join(["?"] * len(selected_processes))
        process_ors = []
        for col in process_columns:
            process_ors.append(f"{escape_sql_identifier(col)} IN ({placeholders})")
            params.extend(selected_processes)
        conditions.append("(" + " OR ".join(process_ors) + ")")

    if date_column and date_start and date_end and date_column in db_columns:
        escaped_date_col = escape_sql_identifier(date_column)
        normalized_date_expr = (
            "CASE "
            f"WHEN LTRIM(RTRIM(CONVERT(NVARCHAR(255), {escaped_date_col}))) IN ('', '0', '0.0') THEN NULL "
            f"ELSE TRY_CONVERT(datetime, {escaped_date_col}) "
            "END"
        )
        conditions.append(f"{normalized_date_expr} >= TRY_CONVERT(datetime, ?)")
        params.append(date_start)
        conditions.append(f"{normalized_date_expr} < DATEADD(day, 1, TRY_CONVERT(datetime, ?))")
        params.append(date_end)

    if column_filters:
        try:
            parsed_filters = json.loads(column_filters)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"列筛选参数格式错误: {exc}")

        if isinstance(parsed_filters, dict):
            for col_name, filter_value in parsed_filters.items():
                if col_name not in db_columns:
                    continue
                text = str(filter_value or "").strip()
                if not text:
                    continue
                escaped_col = escape_sql_identifier(col_name)
                conditions.append(f"LTRIM(RTRIM(CONVERT(NVARCHAR(255), {escaped_col}))) = ?")
                params.append(text)

    where_sql = ""
    if conditions:
        where_sql = " WHERE " + " AND ".join(conditions)

    return where_sql, params


def build_report_query(
    item_code: Optional[str],
    product_name: Optional[str],
    product_spec: Optional[str],
    db_columns,
    workshop_values: Optional[str] = None,
    process_values: Optional[str] = None,
    date_column: Optional[str] = None,
    date_start: Optional[str] = None,
    date_end: Optional[str] = None,
    column_filters: Optional[str] = None,
    offset: Optional[int] = None,
    limit: Optional[int] = None,
):
    where_sql, params = build_report_where(
        item_code,
        product_name,
        product_spec,
        db_columns,
        workshop_values,
        process_values,
        date_column,
        date_start,
        date_end,
        column_filters,
    )

    query = """
        SELECT *
        FROM [APS_SUO].[dbo].[报工统计_开料_锁体A_sw]
    """ + where_sql

    # 默认按第1道/开料尾工序时间升序，无该列时回退到首列。
    first_process_time_col = next(
        (
            col
            for col in db_columns
            if (
                re.search(r"第\s*1\s*道工序", str(col))
                or ("开料尾工序" in str(col))
            ) and re.search(r"时间|日期", str(col))
        ),
        None,
    )
    sort_col = first_process_time_col if first_process_time_col else (db_columns[0] if db_columns else "料品编码")
    query += f" ORDER BY {escape_sql_identifier(sort_col)} ASC"

    # 分批拉取时使用稳定排序，确保分页结果可预期。
    if offset is not None and limit is not None:
        query += " OFFSET ? ROWS FETCH NEXT ? ROWS ONLY"
        params.extend([int(offset), int(limit)])

    return query, params


def fetch_report_total(
    cursor,
    item_code: Optional[str],
    product_name: Optional[str],
    product_spec: Optional[str],
    db_columns,
    workshop_values: Optional[str] = None,
    process_values: Optional[str] = None,
    date_column: Optional[str] = None,
    date_start: Optional[str] = None,
    date_end: Optional[str] = None,
    column_filters: Optional[str] = None,
):
    where_sql, params = build_report_where(
        item_code,
        product_name,
        product_spec,
        db_columns,
        workshop_values,
        process_values,
        date_column,
        date_start,
        date_end,
        column_filters,
    )
    count_query = "SELECT COUNT(1) FROM [APS_SUO].[dbo].[报工统计_开料_锁体A_sw]" + where_sql
    cursor.execute(count_query, params)
    return int(cursor.fetchone()[0])


def rename_process_columns(columns: list, data: list):
    """重命名工序列，添加报工总数相差列，以及插入新的空列"""
    if not data:
        return columns, data

    def parse_to_date(value):
        if value is None:
            return None

        if isinstance(value, datetime):
            return value.date()

        text = str(value).strip()
        if not text or text == "-":
            return None

        # 兼容 ISO 时间与常见日期格式
        iso_text = text.replace("Z", "+00:00")
        try:
            return datetime.fromisoformat(iso_text).date()
        except Exception:
            pass

        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d",
            "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d %H:%M",
            "%Y/%m/%d",
        ):
            try:
                return datetime.strptime(text, fmt).date()
            except Exception:
                continue

        return None
    
    # 列名映射规则
    mapping = {}
    proc1_report_col = None
    proc2_report_col = None
    
    for col in columns:
        if "第1道工序" in col:
            mapping[col] = col.replace("第1道工序", "开料尾工序")
            if "报工数" in col:
                proc1_report_col = col
        elif "第2道工序" in col:
            mapping[col] = col.replace("第2道工序", "锁体A首工序")
            if "报工数" in col:
                proc2_report_col = col
    
    # 转换原列名（避免重名覆盖：若目标名已存在，则保留原列名）
    existing_original = set(columns)
    mapped_columns = []
    for col in columns:
        target = mapping.get(col, col)
        if target != col and target in existing_original:
            # 目标列名在原始列已存在，跳过改名以避免同名冲突
            mapped_columns.append(col)
        else:
            mapped_columns.append(target)

    new_columns = mapped_columns

    # old_col -> final_col 映射，后续构造数据时保持和 new_columns 一致
    final_name_map = {old: new for old, new in zip(columns, new_columns)}
    
    # 找到需要插入新列的位置 (按倒序以避免位置偏移)
    insert_positions = []  # (位置, 新列名)
    
    # 1. 在"锁体A首工序最早结束时间"后面插入"锁体A首工序最晚结束时间"
    for i, col in enumerate(new_columns):
        if "锁体A首工序" in col and "最早结束时间" in col:
            insert_positions.append((i + 1, "锁体A首工序最晚结束时间"))
            break
    
    # 2. 在"开料尾工序最早结束时间"后面插入"开料尾工序最晚结束时间"
    for i, col in enumerate(new_columns):
        if "开料尾工序" in col and "最早结束时间" in col:
            insert_positions.append((i + 1, "开料尾工序最晚结束时间"))
            break
    
    # 3. 在最后一个"开料尾工序"相关列（或新插入的最晚结束时间）后面找"涉及车间"，在其前面插入"距离今天天数"
    proc1_latest_idx = -1
    for i, col in enumerate(new_columns):
        if "开料尾工序" in col:
            proc1_latest_idx = i
    
    workshop_idx = -1
    for i, col in enumerate(new_columns):
        if "所涉及车间" in col:
            workshop_idx = i
            break
    
    if proc1_latest_idx >= 0:
        if workshop_idx >= 0 and workshop_idx > proc1_latest_idx:
            # 在开料尾工序之后、涉及车间之前插入
            insert_positions.append((proc1_latest_idx + 1, "距离今天天数"))
        else:
            # 如果没有涉及车间，就在开料尾工序最后一列之后插入
            insert_positions.append((proc1_latest_idx + 1, "距离今天天数"))
    
    # 避免重复插入同名新增列（导入过已转换Excel时会出现）
    insert_positions = [
        (pos, col_name)
        for pos, col_name in insert_positions
        if col_name not in new_columns
    ]

    # 按倒序插入以避免位置偏移
    insert_positions.sort(reverse=True)
    for pos, col_name in insert_positions:
        new_columns.insert(pos, col_name)
    
    # 如果有报工数列，添加计算列"报工总数相差"
    if proc1_report_col and proc2_report_col and "报工总数相差" not in new_columns:
        new_columns.append("报工总数相差")

    # 在最后追加列
    day_distance_columns = [
        "开料尾工序最晚结束时间距离今天的时间",
        "锁体A首工序最晚结束时间距离今天的时间",
        "开料-锁体A时间差",
    ]
    for col_name in day_distance_columns:
        if col_name not in new_columns:
            new_columns.append(col_name)
    
    # 转换数据
    new_data = []
    for row in data:
        try:
            # 先按原列名处理
            new_row = {}
            for old_col in columns:
                new_col = final_name_map.get(old_col, old_col)
                # 使用.get()避免KeyError
                new_row[new_col] = row.get(old_col)
            
            # 添加新的空列
            for _, col_name in reversed(insert_positions):
                new_row[col_name] = None
            
            # 添加计算列
            if proc1_report_col and proc2_report_col:
                val1 = row.get(proc1_report_col)
                val2 = row.get(proc2_report_col)
                # 转换为数字后做差
                try:
                    num1 = int(val1) if val1 is not None else 0
                    num2 = int(val2) if val2 is not None else 0
                    new_row["报工总数相差"] = num2 - num1
                except (ValueError, TypeError):
                    new_row["报工总数相差"] = None

            # 计算两个“最晚结束时间”到今天的天数距离（绝对天数）
            today = datetime.now().date()

            proc1_latest_date = parse_to_date(new_row.get("开料尾工序最晚结束时间"))
            if proc1_latest_date is None:
                new_row["开料尾工序最晚结束时间距离今天的时间"] = None
            else:
                new_row["开料尾工序最晚结束时间距离今天的时间"] = abs((proc1_latest_date - today).days)

            proc2_latest_date = parse_to_date(new_row.get("锁体A首工序最晚结束时间"))
            if proc2_latest_date is None:
                new_row["锁体A首工序最晚结束时间距离今天的时间"] = None
            else:
                new_row["锁体A首工序最晚结束时间距离今天的时间"] = abs((proc2_latest_date - today).days)

            # 新增：锁体A首工序最晚结束时间 - 开料尾工序最晚结束时间（单位：天）
            if proc1_latest_date is None or proc2_latest_date is None:
                new_row["开料-锁体A时间差"] = None
            else:
                new_row["开料-锁体A时间差"] = (proc2_latest_date - proc1_latest_date).days
            
            new_data.append(new_row)
        except Exception as e:
            # 如果转换出错，至少返回原始行的数据
            print(f"行转换错误: {e}, 原始列: {columns}, 原始行: {row}")
            new_data.append(row)
    
    return new_columns, new_data


def fetch_report_rows(
    cursor,
    item_code: Optional[str],
    product_name: Optional[str],
    product_spec: Optional[str],
    db_columns,
    workshop_values: Optional[str] = None,
    process_values: Optional[str] = None,
    date_column: Optional[str] = None,
    date_start: Optional[str] = None,
    date_end: Optional[str] = None,
    column_filters: Optional[str] = None,
    offset: Optional[int] = None,
    limit: Optional[int] = None,
):
    query, params = build_report_query(
        item_code,
        product_name,
        product_spec,
        db_columns,
        workshop_values,
        process_values,
        date_column,
        date_start,
        date_end,
        column_filters,
        offset,
        limit,
    )
    cursor.execute(query, params)
    columns = [column[0] for column in cursor.description]
    rows = cursor.fetchall()

    def normalize_datetime_placeholder(col_name, value):
        name = str(col_name or "")
        if not re.search(r"时间|日期", name):
            return value

        if value in (0, 0.0, "0", "0.0"):
            return None

        if isinstance(value, str) and value.strip() in ("0", "0.0"):
            return None

        return value

    data = []
    for row in rows:
        row_dict = {}
        for i, value in enumerate(row):
            value = normalize_datetime_placeholder(columns[i], value)
            if hasattr(value, "isoformat"):
                row_dict[columns[i]] = value.isoformat()
            else:
                row_dict[columns[i]] = value
        data.append(row_dict)

    # 重命名工序列并添加计算列
    columns, data = rename_process_columns(columns, data)

    return columns, data


def get_table_columns(cursor):
    cursor.execute(
        """
        SELECT COLUMN_NAME
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = 'dbo' AND TABLE_NAME = '报工统计_开料_锁体A_sw'
        ORDER BY ORDINAL_POSITION
        """
    )
    return [row[0] for row in cursor.fetchall()]


def escape_sql_identifier(name: str) -> str:
    return f"[{name.replace(']', ']]')}]"


def add_missing_columns(cursor, headers, db_columns):
    db_column_set = set(db_columns)
    missing_columns = [col for col in headers if col not in db_column_set]

    for col in missing_columns:
        # SQL Server 列名最长 128 字符
        if len(col) > 128:
            raise HTTPException(status_code=400, detail=f"列名过长(>128): {col}")
        # 动态新增列，使用 NVARCHAR(MAX) 避免长文本导入失败
        cursor.execute(
            f"ALTER TABLE [APS_SUO].[dbo].[报工统计_开料_锁体A_sw] ADD {escape_sql_identifier(col)} NVARCHAR(MAX) NULL"
        )

    return missing_columns


def drop_extra_columns(cursor, headers, db_columns):
    header_set = set(headers)
    extra_columns = [col for col in db_columns if col not in header_set]

    for col in extra_columns:
        cursor.execute(
            f"ALTER TABLE [APS_SUO].[dbo].[报工统计_开料_锁体A_sw] DROP COLUMN {escape_sql_identifier(col)}"
        )

    return extra_columns


def sync_table_columns_to_headers(cursor, headers):
    db_columns = get_table_columns(cursor)
    dropped_columns = drop_extra_columns(cursor, headers, db_columns)
    # 重新读取列，确保后续新增基于最新结构
    latest_columns = get_table_columns(cursor)
    added_columns = add_missing_columns(cursor, headers, latest_columns)
    return added_columns, dropped_columns


def parse_excel_rows(file_content: bytes):
    workbook = load_workbook(filename=io.BytesIO(file_content), data_only=False)
    sheet = workbook.active

    if sheet.max_row < 1 or sheet.max_column < 1:
        raise HTTPException(status_code=400, detail="Excel为空，请上传包含表头和数据的文件")

    def normalize_header(value):
        if value is None:
            return ""
        text = str(value)
        # 兼容 Excel 中不可见空白字符，避免列名识别异常
        text = text.replace("\u3000", " ").replace("\xa0", " ").replace("\r", " ").replace("\n", " ")
        text = re.sub(r"\s+", " ", text)
        return text.strip()

    headers = []
    header_col_indexes = []
    for col_idx in range(1, sheet.max_column + 1):
        header_value = normalize_header(sheet.cell(row=1, column=col_idx).value)
        if header_value:
            headers.append(header_value)
            header_col_indexes.append(col_idx)

    if not headers:
        raise HTTPException(status_code=400, detail="Excel表头为空")

    if len(set(headers)) != len(headers):
        raise HTTPException(status_code=400, detail="表头存在重复列名，请修正后重试")

    data_rows = []
    skipped_rows = 0
    for row_idx in range(2, sheet.max_row + 1):
        row_values = []
        has_non_empty_value = False

        for col_idx in header_col_indexes:
            value = sheet.cell(row=row_idx, column=col_idx).value
            # 不在这里strip字符串，保待给clean_import_data处理
            if value is not None and value != "":
                has_non_empty_value = True
            row_values.append(value)

        # 跳过整行空数据
        if not has_non_empty_value:
            skipped_rows += 1
            continue

        data_rows.append(row_values)

    if not data_rows:
        raise HTTPException(status_code=400, detail=f"Excel没有可导入的数据行（表头行后全为空）")

    return headers, data_rows


@router.get("/report/raw", summary="获取原始报表数据（调试用）", description="不进行任何转换，返回数据库原始数据")
async def get_report_raw_data(
    limit: Optional[int] = Query(default=10, ge=1, le=100, description="返回行数限制"),
):
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 直接查询，不进行任何转换
        query = f"SELECT TOP {limit} * FROM [APS_SUO].[dbo].[报工统计_开料_锁体A_sw]"
        cursor.execute(query)
        columns = [column[0] for column in cursor.description]
        rows = cursor.fetchall()

        data = []
        for row in rows:
            row_dict = {}
            for i, value in enumerate(row):
                if hasattr(value, "isoformat"):
                    row_dict[columns[i]] = value.isoformat()
                else:
                    row_dict[columns[i]] = value
            data.append(row_dict)

        return {
            "status": "success",
            "columns": columns,
            "data": data,
            "row_count": len(data),
            "timestamp": datetime.now().isoformat(),
        }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取原始数据失败: {exc}",
        )
    finally:
        if conn:
            conn.close()


@router.get("/report", summary="获取盘点报表", description="查询 APS_SUO.dbo.报工统计_开料_锁体A_sw 并支持前3列筛选")
async def get_report_data(
    item_code: Optional[str] = None,
    product_name: Optional[str] = None,
    product_spec: Optional[str] = None,
    workshop_values: Optional[str] = Query(default=None, description="逗号分隔车间值"),
    process_values: Optional[str] = Query(default=None, description="逗号分隔工序值"),
    date_column: Optional[str] = Query(default=None, description="日期筛选列名"),
    date_start: Optional[str] = Query(default=None, description="日期范围起始 yyyy-MM-dd"),
    date_end: Optional[str] = Query(default=None, description="日期范围结束 yyyy-MM-dd"),
    column_filters: Optional[str] = Query(default=None, description="JSON格式列筛选，如 {'第1道工序':'开料'}"),
    offset: Optional[int] = Query(default=None, ge=0, description="分页偏移量，从0开始"),
    limit: Optional[int] = Query(default=None, ge=1, le=5000, description="分页大小，建议1000-5000"),
):
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        db_columns = get_table_columns(cursor)
        total = fetch_report_total(
            cursor,
            item_code,
            product_name,
            product_spec,
            db_columns,
            workshop_values,
            process_values,
            date_column,
            date_start,
            date_end,
            column_filters,
        )
        columns, data = fetch_report_rows(
            cursor,
            item_code,
            product_name,
            product_spec,
            db_columns,
            workshop_values,
            process_values,
            date_column,
            date_start,
            date_end,
            column_filters,
            offset,
            limit,
        )

        return {
            "status": "success",
            "data": data,
            "columns": columns,
            "total": total,
            "timestamp": datetime.now().isoformat(),
        }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取报表数据失败: {exc}",
        )
    finally:
        if conn:
            conn.close()


@router.get("/report/workshops", summary="获取车间筛选选项", description="获取盘点报表车间去重列表")
async def get_report_workshops(
    item_code: Optional[str] = None,
    product_name: Optional[str] = None,
    product_spec: Optional[str] = None,
):
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        db_columns = get_table_columns(cursor)
        workshop_column = get_workshop_column(db_columns)
        if not workshop_column:
            return {
                "status": "success",
                "data": [],
                "total": 0,
            }

        # 车间选项始终返回全量，不受当前筛选条件影响。
        conditions = []
        params = []
        workshop_col_sql = escape_sql_identifier(workshop_column)
        conditions.append(f"{workshop_col_sql} IS NOT NULL")
        conditions.append(f"LTRIM(RTRIM(CONVERT(NVARCHAR(255), {workshop_col_sql}))) <> ''")

        where_sql = " WHERE " + " AND ".join(conditions) if conditions else ""
        query = (
            "SELECT DISTINCT "
            f"CONVERT(NVARCHAR(255), {workshop_col_sql}) AS workshop "
            "FROM [APS_SUO].[dbo].[报工统计_开料_锁体A_sw]"
            f"{where_sql} "
            "ORDER BY workshop"
        )

        cursor.execute(query, params)
        workshops = [str(row[0]).strip() for row in cursor.fetchall() if row[0] is not None and str(row[0]).strip()]

        return {
            "status": "success",
            "data": workshops,
            "total": len(workshops),
        }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取车间选项失败: {exc}",
        )
    finally:
        if conn:
            conn.close()


@router.get("/report/filter-options", summary="获取工序和日期筛选选项", description="返回工序值列表和日期列名")
async def get_report_filter_options(
    item_code: Optional[str] = None,
    product_name: Optional[str] = None,
    product_spec: Optional[str] = None,
    workshop_values: Optional[str] = Query(default=None, description="逗号分隔车间值"),
):
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        db_columns = get_table_columns(cursor)
        process_columns = get_process_name_columns(db_columns)
        date_columns = [
            col
            for col in db_columns
            if (
                re.search(r"第\s*\d+\s*道工序", str(col))
                or ("开料尾工序" in str(col))
                or ("锁体A首工序" in str(col))
            ) and re.search(r"时间|日期", str(col))
        ]

        where_sql, params = build_report_where(
            item_code,
            product_name,
            product_spec,
            db_columns,
            workshop_values=workshop_values,
        )

        process_values = []
        if process_columns:
            table_name = "[APS_SUO].[dbo].[报工统计_开料_锁体A_sw]"
            sub_queries = []
            sub_params = []
            for col in process_columns:
                col_sql = escape_sql_identifier(col)
                null_filter = f"{col_sql} IS NOT NULL AND LTRIM(RTRIM(CONVERT(NVARCHAR(255), {col_sql}))) <> ''"
                if where_sql:
                    sub_query = (
                        "SELECT LTRIM(RTRIM(CONVERT(NVARCHAR(255), " + col_sql + "))) AS process_name "
                        "FROM " + table_name + where_sql + " AND " + null_filter
                    )
                else:
                    sub_query = (
                        "SELECT LTRIM(RTRIM(CONVERT(NVARCHAR(255), " + col_sql + "))) AS process_name "
                        "FROM " + table_name + " WHERE " + null_filter
                    )
                sub_queries.append(sub_query)
                sub_params.extend(params)

            union_sql = " UNION ALL ".join(sub_queries)
            query = "SELECT DISTINCT process_name FROM (" + union_sql + ") AS t ORDER BY process_name"
            cursor.execute(query, sub_params)
            process_values = [
                str(row[0]).strip()
                for row in cursor.fetchall()
                if row[0] is not None and str(row[0]).strip()
            ]

        return {
            "status": "success",
            "process_columns": process_columns,
            "process_options": process_values,
            "date_columns": date_columns,
        }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取工序和日期筛选选项失败: {exc}",
        )
    finally:
        if conn:
            conn.close()


@router.get("/report/column-options", summary="获取指定列筛选选项", description="按当前筛选条件返回指定列去重值")
async def get_report_column_options(
    column_name: str = Query(..., description="列名"),
    item_code: Optional[str] = None,
    product_name: Optional[str] = None,
    product_spec: Optional[str] = None,
    workshop_values: Optional[str] = Query(default=None, description="逗号分隔车间值"),
    process_values: Optional[str] = Query(default=None, description="逗号分隔工序值"),
    date_column: Optional[str] = Query(default=None, description="日期筛选列名"),
    date_start: Optional[str] = Query(default=None, description="日期范围起始 yyyy-MM-dd"),
    date_end: Optional[str] = Query(default=None, description="日期范围结束 yyyy-MM-dd"),
    column_filters: Optional[str] = Query(default=None, description="JSON格式列筛选"),
):
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        db_columns = get_table_columns(cursor)

        if column_name not in db_columns:
            raise HTTPException(status_code=400, detail=f"无效列名: {column_name}")

        where_sql, params = build_report_where(
            item_code,
            product_name,
            product_spec,
            db_columns,
            workshop_values,
            process_values,
            date_column,
            date_start,
            date_end,
            column_filters,
        )

        col_sql = escape_sql_identifier(column_name)
        query = (
            "SELECT DISTINCT LTRIM(RTRIM(CONVERT(NVARCHAR(255), " + col_sql + "))) AS opt "
            "FROM [APS_SUO].[dbo].[报工统计_开料_锁体A_sw]" + where_sql +
            (" AND " if where_sql else " WHERE ") +
            col_sql + " IS NOT NULL AND LTRIM(RTRIM(CONVERT(NVARCHAR(255), " + col_sql + "))) <> '' "
            "ORDER BY opt"
        )

        cursor.execute(query, params)
        options = [str(row[0]).strip() for row in cursor.fetchall() if row[0] is not None and str(row[0]).strip()]

        return {
            "status": "success",
            "data": options,
            "total": len(options),
        }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取列筛选选项失败: {exc}",
        )
    finally:
        if conn:
            conn.close()


@router.get("/report/export", summary="导出盘点报表Excel", description="按筛选条件导出 APS_SUO.dbo.报工统计_开料_锁体A_sw")
async def export_report_excel(
    item_code: Optional[str] = None,
    product_name: Optional[str] = None,
    product_spec: Optional[str] = None,
    workshop_values: Optional[str] = Query(default=None, description="逗号分隔车间值"),
    process_values: Optional[str] = Query(default=None, description="逗号分隔工序值"),
    date_column: Optional[str] = Query(default=None, description="日期筛选列名"),
    date_start: Optional[str] = Query(default=None, description="日期范围起始 yyyy-MM-dd"),
    date_end: Optional[str] = Query(default=None, description="日期范围结束 yyyy-MM-dd"),
    column_filters: Optional[str] = Query(default=None, description="JSON格式列筛选，如 {'第1道工序':'开料'}"),
):
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        db_columns = get_table_columns(cursor)
        columns, data = fetch_report_rows(
            cursor,
            item_code,
            product_name,
            product_spec,
            db_columns,
            workshop_values,
            process_values,
            date_column,
            date_start,
            date_end,
            column_filters,
        )

        if not data:
            raise HTTPException(status_code=404, detail="没有可导出的数据")

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "盘点报表"

        sheet.append(columns)
        for row in data:
            sheet.append([row.get(col) for col in columns])

        for idx, col_name in enumerate(columns, start=1):
            max_len = len(str(col_name))
            for row in data:
                value = row.get(col_name)
                length = len(str(value)) if value is not None else 0
                if length > max_len:
                    max_len = length
            sheet.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 60)

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"盘点报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        encoded_filename = urllib.parse.quote(filename)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
        )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"导出Excel失败: {exc}",
        )
    finally:
        if conn:
            conn.close()


def clean_import_data(data_rows: list):
    """清理导入数据：空字符串转None，处理特殊值"""
    cleaned_rows = []
    for row in data_rows:
        cleaned_row = []
        for value in row:
            # 空字符串或仅空白转换为None
            if isinstance(value, str):
                trimmed = value.strip()
                if not trimmed:
                    cleaned_row.append(None)
                else:
                    cleaned_row.append(trimmed)
            else:
                cleaned_row.append(value)
        cleaned_rows.append(cleaned_row)
    return cleaned_rows


@router.post("/report/import", summary="导入盘点报表Excel", description="上传Excel并覆盖 APS_SUO.dbo.报工统计_开料_锁体A_sw 数据")
async def import_report_excel(file: UploadFile = File(...)):
    conn = None
    try:
        if not file.filename:
            raise HTTPException(status_code=400, detail="未检测到文件名")

        lower_name = file.filename.lower()
        if not (lower_name.endswith(".xlsx") or lower_name.endswith(".xlsm") or lower_name.endswith(".xltx") or lower_name.endswith(".xltm")):
            raise HTTPException(status_code=400, detail="仅支持 .xlsx/.xlsm 格式文件")

        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="上传文件为空")

        headers, data_rows = parse_excel_rows(content)
        
        # 清理数据：空字符串转None，字符串去除空白
        data_rows = clean_import_data(data_rows)

        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 先获取当前数据库中的列
        db_columns = get_table_columns(cursor)
        
        # 检查Excel中的列是否都存在于数据库，如果不存在则新增
        added_columns = []
        for col in headers:
            if col not in db_columns:
                # SQL Server 列名最长 128 字符
                if len(col) > 128:
                    raise HTTPException(status_code=400, detail=f"列名过长(>128): {col}")
                # 动态新增列，使用 NVARCHAR(MAX) 避免长文本导入失败
                cursor.execute(
                    f"ALTER TABLE [APS_SUO].[dbo].[报工统计_开料_锁体A_sw] ADD {escape_sql_identifier(col)} NVARCHAR(MAX) NULL"
                )
                added_columns.append(col)
        
        # 构建INSERT语句
        escaped_columns = [escape_sql_identifier(col) for col in headers]
        placeholders = ", ".join(["?"] * len(headers))
        insert_sql = (
            f"INSERT INTO [APS_SUO].[dbo].[报工统计_开料_锁体A_sw] ({', '.join(escaped_columns)}) "
            f"VALUES ({placeholders})"
        )

        # 清空数据表
        cursor.execute("DELETE FROM [APS_SUO].[dbo].[报工统计_开料_锁体A_sw]")
        
        # 插入数据
        cursor.fast_executemany = True
        
        try:
            if data_rows:
                cursor.executemany(insert_sql, data_rows)
            conn.commit()
        except Exception as insert_exc:
            conn.rollback()
            raise HTTPException(
                status_code=400,
                detail=f"数据插入失败: {insert_exc}。列数: {len(headers)}, 数据行数: {len(data_rows)}。第一行示例: {data_rows[0] if data_rows else '无数据'}"
            )

        # 导入成功后回查数据库行数
        cursor.execute("SELECT COUNT(1) FROM [APS_SUO].[dbo].[报工统计_开料_锁体A_sw]")
        db_total = cursor.fetchone()[0]

        return {
            "status": "success",
            "message": "导入成功",
            "excel_rows": len(headers),
            "inserted_rows": len(data_rows),
            "db_total": db_total,
            "columns": headers,
            "added_columns": added_columns,
            "timestamp": datetime.now().isoformat(),
        }
    except HTTPException:
        if conn:
            conn.rollback()
        raise
    except Exception as exc:
        if conn:
            conn.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"导入Excel失败: {exc}",
        )
    finally:
        if conn:
            conn.close()
